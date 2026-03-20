"""
agent/bom_parser.py
--------------------
Reads BOM Excel → produces a clean service list for the LLM layout compiler.
The LLM receives a structured summary and returns a layout spec JSON.
This file does NOT decide layout — only service identification.
"""
from __future__ import annotations
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ── SKU → (oci_type, default_layer) ─────────────────────────────────────────
SKU_MAP: dict[str, tuple] = {
    "B94176": ("compute",           "compute"),
    "B94177": (None,                None),       # memory — part of compute
    "B91961": (None,                None),       # block storage — implied
    "B91962": (None,                None),       # block perf — implied
    "B99060": ("database",          "data"),
    "B99062": (None,                None),       # db storage — implied
    "B91628": ("object storage",    "data"),
    "B93030": ("load balancer",     "ingress"),
    "B93031": (None,                None),       # LB bandwidth — implied
    "B88325": ("drg",               "ingress"),
    "B90618": ("functions",         "compute"),
    "B90617": (None,                None),       # fn execution — part of functions
    "B92072": ("api gateway",       "ingress"),
    "B95697": ("queue",             "async"),
}

DESC_MAP: dict[str, tuple] = {
    "container instances": ("container engine", "compute"),
    "bastion":             ("bastion",           "ingress"),
    "identity and access": ("iam",               "data"),
    "network load balancer":("load balancer",    "ingress"),
    "secrets on oci vault": ("vault",            "data"),
}

# Best-practice additions (always injected)
BEST_PRACTICE = [
    {"id": "internet_gateway", "type": "internet gateway", "label": "Internet Gateway",     "layer": "external"},
    {"id": "nat_gateway",      "type": "nat gateway",      "label": "NAT Gateway",           "layer": "ingress"},
    {"id": "service_gateway",  "type": "service gateway",  "label": "Service Gateway",       "layer": "ingress"},
    {"id": "waf",              "type": "waf",              "label": "WAF",                   "layer": "ingress"},
    {"id": "network_firewall", "type": "waf",              "label": "Network Firewall",      "layer": "ingress"},
    {"id": "logging",          "type": "logging",          "label": "Logging Analytics\n(SIEM)", "layer": "data"},
    {"id": "monitoring",       "type": "monitoring",       "label": "Monitoring + APM",      "layer": "data"},
    {"id": "db_mgmt",          "type": "monitoring",       "label": "DB Management",         "layer": "data"},
    {"id": "directory",        "type": "iam",              "label": "Directory Services",    "layer": "data"},
    {"id": "certificates",     "type": "vault",            "label": "Certificates",          "layer": "data"},
]


@dataclass
class ServiceItem:
    id:       str
    oci_type: str
    label:    str
    layer:    str           # external | ingress | compute | async | data
    quantity: Optional[float] = None
    notes:    str = ""


def parse_bom(xlsx_path: str | Path) -> list[ServiceItem]:
    """Parse BOM Excel → list of ServiceItems ready for LLM layout prompt."""
    import openpyxl
    wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
    bom  = wb["BOM"]
    inp  = wb["Input"] if "Input" in wb.sheetnames else None

    # Read input sheet for quantities
    input_data: dict[str, dict] = {}
    if inp:
        hdrs = None
        for row in inp.iter_rows(values_only=True):
            if hdrs is None:
                hdrs = [str(c).lower() if c else "" for c in row]
                continue
            if row[0]:
                input_data[str(row[0]).lower()] = {
                    hdrs[i]: row[i] for i in range(1, len(row)) if i < len(hdrs)
                }

    app_ocpu = int((input_data.get("ec2", {}).get("vcpu count", 0) or 0) / 2)
    db_ocpu  = int((input_data.get("postgres rds", {}).get("vcpu count", 0) or 0) / 2)
    obj_gb   = input_data.get("postgres rds", {}).get("storage (gb)", 0) or 0

    items: list[ServiceItem] = []
    seen_types: set[str] = set()
    counters: dict[str, int] = {}

    hdrs = None
    for row in bom.iter_rows(values_only=True):
        if hdrs is None:
            hdrs = [str(c).lower().strip() if c else f"c{i}" for i, c in enumerate(row)]
            continue
        if not any(v is not None for v in row):
            continue
        d = dict(zip(hdrs, row))

        sku  = str(d.get("sku", "") or "").strip()
        desc = str(d.get("description", "") or "").lower().strip()
        qty  = d.get("quantity")
        note = str(d.get(hdrs[-1], "") or "")

        # SKU lookup
        if sku in SKU_MAP:
            oci_type, layer = SKU_MAP[sku]
            if not oci_type or not layer:
                continue
        else:
            # Description lookup
            oci_type, layer = None, None
            for key, (t, l) in DESC_MAP.items():
                if key in desc:
                    oci_type, layer = t, l
                    break
            if not oci_type:
                continue

        # Deduplicate by oci_type (one icon per service type)
        if oci_type in seen_types:
            continue
        seen_types.add(oci_type)

        counters[oci_type] = counters.get(oci_type, 0) + 1
        nid = f"{oci_type.replace(' ', '_')}_{counters[oci_type]}"

        # Build label with quantity context
        label = _make_label(oci_type, qty, app_ocpu, db_ocpu, obj_gb, note)

        items.append(ServiceItem(id=nid, oci_type=oci_type, label=label,
                                 layer=layer, quantity=qty, notes=note))

    # Add On-Premises (always — FastConnect implies it)
    items.insert(0, ServiceItem(id="on_prem", oci_type="on premises",
                                label="On-Premises\n(3 Offices)", layer="external"))

    # Add best-practice services
    for bp in BEST_PRACTICE:
        if bp["type"] not in seen_types:
            items.append(ServiceItem(id=bp["id"], oci_type=bp["type"],
                                     label=bp["label"], layer=bp["layer"],
                                     notes="best practice"))
            seen_types.add(bp["type"])

    return items


def _make_label(oci_type: str, qty, app_ocpu: int, db_ocpu: int, obj_gb: float, note: str) -> str:
    labels = {
        "compute":      f"Compute\n×{app_ocpu:,} OCPU",
        "database":     f"PostgreSQL DB\n×{db_ocpu:,} OCPU",
        "object storage": f"Object Storage\n{int(obj_gb*2/1024)} TB",
        "load balancer": f"Load Balancer\n×{int(qty) if qty else '?'} (per region)",
        "drg":          f"DRG / FastConnect\n×{int(qty) if qty else '?'} ports",
        "functions":    "OCI Functions\n~10k calls/day",
        "api gateway":  "API Gateway",
        "queue":        f"Queue\n{int(qty) if qty else '?'}M req/month",
        "container engine": "Container Instances",
        "bastion":      "Bastion",
        "iam":          "IAM",
        "load balancer_2": f"NLB ×3",
        "vault":        "Vault (Secrets)",
    }
    return labels.get(oci_type, oci_type.title())


def build_llm_prompt(items: list[ServiceItem], context: str = "") -> str:
    """Build the layout compiler prompt for the OCI GenAI agent.

    context: optional text from a requirements/notes file uploaded alongside the BOM.
    If key info is missing the LLM should ask clarification questions rather than guess.
    """
    service_list = "\n".join(
        f'  {{"id": "{i.id}", "type": "{i.oci_type}", "label": "{i.label.replace(chr(10), " ")}", "suggested_layer": "{i.layer}"}}'
        for i in items
    )

    context_block = (
        f"\nADDITIONAL CONTEXT:\n{context.strip()}\n"
        if context and context.strip() else ""
    )

    return f"""You are a layout compiler. Your job is to produce a deterministic layout specification JSON.
{context_block}
CLARIFICATION RULE:
If you are missing information that would materially change the diagram topology
(e.g. number of regions, HA pattern, dedicated subnets), return ONLY this JSON:
{{
  "status": "need_clarification",
  "questions": ["Question 1?", "Question 2?"]
}}
Only ask questions you cannot answer from the BOM or context above.
If you can produce a representative diagram, skip questions and output the spec.

INPUT SERVICES (from BOM):
[
{service_list}
]

LAYER ORDER (strict left-to-right):
  1. external   — outside OCI (on-premises, internet)
  2. ingress    — entry points (gateways, LBs, WAF, Firewall, Bastion, DRG)
  3. compute    — processing (VMs, containers, functions)
  4. async      — messaging (queues, streaming)
  5. data       — storage and databases

RULES:
1. Every node MUST be assigned to exactly one layer using the order above.
2. Groups MUST use OCI subnet names: "Public Subnet", "App Subnet", "DB Subnet", "OCI Region Services"
3. Gateways (internet_gateway, nat_gateway, service_gateway, drg) go in "Public Subnet" group at ingress layer.
4. Security (waf, network_firewall, bastion) go in "Public Subnet" group at ingress layer.
5. Load balancers go in "Public Subnet" group at ingress layer.
6. Compute, functions, containers, api_gateway go in "App Subnet" group at compute layer.
7. Queue goes in "App Subnet" group at async layer.
8. Database, vault go in "DB Subnet" group at data layer.
9. Object storage, IAM, logging, monitoring, certificates, directory, db_mgmt go in "OCI Region Services" group at data layer.
10. on_prem is NOT in any group — it is external.
11. Edges connect layers left-to-right. No backward edges.
12. Use these edges ONLY — no extras:
    - on_prem → drg_1 (FastConnect ×6)
    - internet_gateway → pub_sub_box (Internet)
    - drg_1 → app_sub_box (internal routing)
    - pub_sub_box → app_sub_box (LB Traffic)
    - bastion → app_sub_box (SSH / Admin)
    - app_sub_box → db_sub_box (Data Access)
    - service_gateway → object_storage_1 (OCI Backbone)

OUTPUT this exact JSON structure (fill in the nodes arrays):
{{
  "direction": "LR",
  "page": {{"width": 1654, "height": 1169}},
  "layers": {{
    "external": [],
    "ingress":  [],
    "compute":  [],
    "async":    [],
    "data":     []
  }},
  "groups": [
    {{"id": "pub_sub_box",  "label": "Public Subnet",         "nodes": []}},
    {{"id": "app_sub_box",  "label": "App Subnet",            "nodes": []}},
    {{"id": "db_sub_box",   "label": "DB Subnet",             "nodes": []}},
    {{"id": "region_box",   "label": "OCI Region Services",   "nodes": []}}
  ],
  "edges": [
    {{"id": "e1", "source": "on_prem",          "target": "drg_1",        "label": "FastConnect ×6"}},
    {{"id": "e2", "source": "internet_gateway", "target": "pub_sub_box",  "label": "Internet"}},
    {{"id": "e3", "source": "drg_1",            "target": "app_sub_box",  "label": ""}},
    {{"id": "e4", "source": "pub_sub_box",      "target": "app_sub_box",  "label": "LB Traffic"}},
    {{"id": "e5", "source": "bastion",          "target": "app_sub_box",  "label": "SSH / Admin"}},
    {{"id": "e6", "source": "app_sub_box",      "target": "db_sub_box",   "label": "Data Access"}},
    {{"id": "e7", "source": "service_gateway",  "target": "object_storage_1", "label": ""}}
  ]
}}

Fill ONLY the nodes arrays in layers and groups. Do not change edges. Output ONLY valid JSON."""


def bom_to_llm_input(xlsx_path: str | Path, context: str = "") -> tuple[list[ServiceItem], str]:
    """Main entry point: parse BOM and return (items, llm_prompt).

    context: optional free-text from an uploaded requirements/notes file.
    """
    items = parse_bom(xlsx_path)
    prompt = build_llm_prompt(items, context=context)
    return items, prompt
