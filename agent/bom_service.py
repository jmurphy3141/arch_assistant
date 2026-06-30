from __future__ import annotations

import io
import json
import logging
import re
import time
import urllib.request
import uuid
from dataclasses import dataclass
from threading import RLock
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Font

from agent import consistency_contract

logger = logging.getLogger(__name__)

OCI_PRICE_LIST_API_URL = "https://apexapps.oracle.com/pls/apex/cetools/api/v1/products/"
OCI_SHAPES_DOC_URL = "https://docs.oracle.com/en-us/iaas/Content/Compute/References/computeshapes.htm"
OCI_SERVICES_DOC_URL = "https://docs.oracle.com/en-us/iaas/Content/home.htm"

CPU_SKU_TO_MEM_SKU = {
    "B93113": "B93114",
    "B97384": "B97385",
    "B111129": "B111130",
    "B94176": "B94177",
    "B93297": "B93298",
}

DEFAULT_PRICE_TABLE: dict[str, dict[str, Any]] = {
    "B93113": {"description": "Compute - Standard - E4 - OCPU", "unit_price": 0.025, "metric": "OCPU Per Hour", "category": "compute"},
    "B93114": {"description": "Compute - Standard - E4 - Memory", "unit_price": 0.0015, "metric": "Gigabyte Per Hour", "category": "compute"},
    "B97384": {"description": "Compute - Standard - E5 - OCPU", "unit_price": 0.03, "metric": "OCPU Per Hour", "category": "compute"},
    "B97385": {"description": "Compute - Standard - E5 - Memory", "unit_price": 0.002, "metric": "Gigabytes Per Hour", "category": "compute"},
    "B111129": {"description": "OCI - Compute - Standard - E6 - OCPU", "unit_price": 0.03, "metric": "OCPU Per Hour", "category": "compute"},
    "B111130": {"description": "OCI - Compute - Standard - E6 - Memory", "unit_price": 0.002, "metric": "Gigabytes Per Hour", "category": "compute"},
    "B94176": {"description": "Compute - Standard - X9 - OCPU", "unit_price": 0.04, "metric": "OCPU Per Hour", "category": "compute"},
    "B94177": {"description": "Compute - Standard - X9 - Memory", "unit_price": 0.0015, "metric": "Gigabytes Per Hour", "category": "compute"},
    "B93297": {"description": "Compute - Standard - A1 - OCPU", "unit_price": 0.0, "metric": "OCPU Per Hour", "category": "compute"},
    "B93298": {"description": "Compute - Standard - A1 - Memory", "unit_price": 0.0, "metric": "Gigabyte Per Hour", "category": "compute"},
    "B91961": {"description": "Storage - Block Volume - Storage", "unit_price": 0.0255, "metric": "Gigabyte Storage Capacity Per Month", "category": "storage"},
    "B91962": {"description": "Storage - Block Volume - Performance Units", "unit_price": 0.0017, "metric": "Performance Unit Per Gigabyte Per Month", "category": "storage"},
    "B93030": {"description": "Load Balancer Base", "unit_price": 0.0113, "metric": "Load Balancer Base Instance Per Hour", "category": "network"},
    "B91628": {"description": "Object Storage - Storage", "unit_price": 0.0255, "metric": "Gigabyte Storage Capacity Per Month", "category": "storage"},
    "B99060": {"description": "Oracle Autonomous Database - ECPU", "unit_price": 0.0672, "metric": "ECPU Per Hour", "category": "database"},
    "B88325": {"description": "FastConnect - 1 Gbps Port Hour", "unit_price": 0.212, "metric": "Port Hour", "category": "network"},
    "BFILE01": {"description": "File Storage - Capacity", "unit_price": 0.0255, "metric": "Gigabyte Storage Capacity Per Month", "category": "storage"},
    "BWAF01": {"description": "Web Application Firewall Policy", "unit_price": 0.6, "metric": "Policy Per Hour", "category": "network"},
}

NON_OCI_PROVIDER_PATTERNS: tuple[str, ...] = (
    "hetzner",
    "digitalocean",
    "do ",
    "vultr",
    "linode",
    "aws",
    "amazon web services",
    "ec2",
    "azure",
    "gcp",
    "google cloud",
    "cloudflare",
)

_CAPACITY_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(tb|gb)\b", re.IGNORECASE)


@dataclass
class CacheSnapshot:
    pricing_table: dict[str, dict[str, Any]]
    shapes_text: str
    services_text: str
    refreshed_at: float
    source: str


class BomService:
    """
    Shared v1.7 BOM service used by REST and orchestrator tool execution.
    """

    def __init__(self) -> None:
        self._lock = RLock()
        self._cache: CacheSnapshot | None = None

    def _fetch_url(self, url: str, timeout: int = 20) -> str:
        req = urllib.request.Request(url, headers={"User-Agent": "oci-agent-bom/1.7"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read().decode("utf-8", errors="replace")

    def refresh_data(self) -> dict[str, Any]:
        started = time.perf_counter()
        pricing_table: dict[str, dict[str, Any]] = {}
        source = "live"

        try:
            payload = self._fetch_pricing_payload()
            pricing_table = self._parse_pricing_payload(payload)
            if not pricing_table:
                raise ValueError("pricing payload parsed to empty table")
        except Exception as exc:
            logger.warning("BOM pricing refresh fallback: %s", exc)
            source = "fallback"
            pricing_table = self._unpriced_default_catalog()

        try:
            shapes_text = self._fetch_url(OCI_SHAPES_DOC_URL)
            services_text = self._fetch_url(OCI_SERVICES_DOC_URL)
        except Exception as exc:
            logger.warning("BOM shapes/services refresh fallback: %s", exc)
            source = "fallback" if source == "fallback" else "live_pricing_partial_reference_fallback"
            shapes_text = "OCI compute shapes catalog unavailable; using fallback guidance."
            services_text = "OCI services catalog unavailable; using fallback guidance."

        with self._lock:
            self._cache = CacheSnapshot(
                pricing_table=pricing_table,
                shapes_text=shapes_text[:60000],
                services_text=services_text[:60000],
                refreshed_at=time.time(),
                source=source,
            )

        elapsed_ms = int((time.perf_counter() - started) * 1000)
        return {
            "ready": True,
            "source": source,
            "pricing_sku_count": len(pricing_table),
            "latency_ms": elapsed_ms,
            "refreshed_at": int(self._cache.refreshed_at) if self._cache else None,
        }

    def _fetch_pricing_payload(self) -> dict[str, Any]:
        items_all: list[dict[str, Any]] = []
        url = OCI_PRICE_LIST_API_URL
        last_updated: Any = None
        seen_urls: set[str] = set()

        while url and url not in seen_urls:
            seen_urls.add(url)
            payload = json.loads(self._fetch_url(url, timeout=30))
            if not isinstance(payload, dict):
                raise ValueError("pricing payload was not a JSON object")
            if last_updated is None:
                last_updated = payload.get("lastUpdated")
            items = payload.get("items") or []
            if isinstance(items, list):
                items_all.extend(item for item in items if isinstance(item, dict))
            next_link = next(
                (
                    str(link.get("href"))
                    for link in (payload.get("links") or [])
                    if isinstance(link, dict) and str(link.get("rel") or "").lower() == "next" and link.get("href")
                ),
                "",
            )
            url = next_link

        return {"lastUpdated": last_updated, "items": items_all}

    def _parse_pricing_payload(self, payload: Any) -> dict[str, dict[str, Any]]:
        table: dict[str, dict[str, Any]] = {}
        rows: list[dict[str, Any]] = []
        if isinstance(payload, list):
            rows = [r for r in payload if isinstance(r, dict)]
        elif isinstance(payload, dict):
            for key in ("items", "products", "data"):
                val = payload.get(key)
                if isinstance(val, list):
                    rows = [r for r in val if isinstance(r, dict)]
                    break

        for row in rows:
            sku = str(row.get("sku") or row.get("partNumber") or "").strip().upper()
            if not sku:
                continue

            description = str(row.get("description") or row.get("displayName") or "").strip()
            unit_price = self._extract_usd_unit_price(row)
            if unit_price is None:
                continue

            category = "compute" if "compute" in description.lower() else "service"
            if any(token in description.lower() for token in ("storage", "volume", "object")):
                category = "storage"
            if any(token in description.lower() for token in ("load balancer", "network", "gateway", "fastconnect")):
                category = "network"
            table[sku] = {
                "description": description or sku,
                "unit_price": unit_price,
                "metric": str(row.get("metric") or row.get("metricName") or "").strip(),
                "category": category,
                "source": "oracle_price_list_api",
            }

        # Keep catalog metadata for omitted SKUs, but never backfill a price
        # that the current Oracle public price response did not provide.
        for sku, row in DEFAULT_PRICE_TABLE.items():
            if sku not in table:
                fallback = dict(row)
                fallback["unit_price"] = None
                fallback["source"] = "unpriced_catalog"
                fallback["pricing_status"] = "unpriced"
                table[sku] = fallback
        return table

    @staticmethod
    def _unpriced_default_catalog() -> dict[str, dict[str, Any]]:
        output: dict[str, dict[str, Any]] = {}
        for sku, row in DEFAULT_PRICE_TABLE.items():
            item = dict(row)
            item["unit_price"] = None
            item["source"] = "unpriced_catalog"
            item["pricing_status"] = "unpriced"
            output[sku] = item
        return output

    @classmethod
    def _extract_usd_unit_price(cls, row: dict[str, Any]) -> float | None:
        for key in ("unit_cost", "unitPrice", "price"):
            value = cls._parse_numeric(row.get(key))
            if value is not None:
                return value

        localizations = row.get("currencyCodeLocalizations") or []
        if not isinstance(localizations, list):
            return None
        usd = next(
            (
                item
                for item in localizations
                if isinstance(item, dict) and str(item.get("currencyCode") or "").upper() == "USD"
            ),
            None,
        )
        if not usd:
            return None
        prices = usd.get("prices") or []
        if not isinstance(prices, list):
            return None
        return cls._select_unit_cost_from_price_tiers([p for p in prices if isinstance(p, dict)])

    @classmethod
    def _select_unit_cost_from_price_tiers(cls, prices: list[dict[str, Any]]) -> float | None:
        if not prices:
            return None

        payg = [p for p in prices if str(p.get("model") or "").upper() == "PAY_AS_YOU_GO"]
        candidates = payg or prices
        positive_tiers: list[tuple[float, int, float]] = []
        for idx, tier in enumerate(candidates):
            value = cls._parse_numeric(tier.get("value"))
            if value is None or value <= 0:
                continue
            range_min = cls._parse_numeric(tier.get("rangeMin"))
            positive_tiers.append((range_min if range_min is not None else 0.0, idx, value))

        if positive_tiers:
            positive_tiers.sort(key=lambda item: (item[0], item[1]))
            return positive_tiers[0][2]

        for tier in candidates:
            value = cls._parse_numeric(tier.get("value"))
            if value is not None:
                return value
        return None

    @staticmethod
    def _parse_numeric(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
        return float(match.group(0)) if match else None

    def health(self) -> dict[str, Any]:
        with self._lock:
            snap = self._cache
        if not snap:
            return {
                "ready": False,
                "source": "none",
                "refreshed_at": None,
                "pricing_sku_count": 0,
            }
        return {
            "ready": True,
            "source": snap.source,
            "refreshed_at": int(snap.refreshed_at),
            "pricing_sku_count": len(snap.pricing_table),
        }

    def config(self, default_model_id: str) -> dict[str, Any]:
        status = self.health()
        return {
            "status": "ok",
            "default_model_id": default_model_id,
            "cache": status,
            "allowed_types": ["normal", "question", "final"],
        }

    def _build_shape_catalog(self, price_table: dict[str, dict]) -> str:
        """
        Build a compact shape→SKU reference from the live price table.
        Returns a pipe-delimited string for injection into LLM prompts.
        """
        SHAPE_KEYWORDS = [
            ("e4", "E4.Flex (AMD)", "OCPU"),
            ("e5", "E5.Flex (AMD)", "OCPU"),
            ("e6", "E6.Flex (AMD)", "OCPU"),
            ("x9", "X9 (Intel Standard)", "OCPU"),
            ("a1", "A1.Flex (Ampere)", "OCPU"),
            ("gpu", "GPU shapes", "GPU"),
        ]
        lines = ["Shape | CPU SKU | CPU $/hr | Mem SKU | Mem $/hr"]
        for keyword, label, _ in SHAPE_KEYWORDS:
            for sku, row in price_table.items():
                desc = row.get("description", "").lower()
                metric = row.get("metric", "").lower()
                if keyword in desc and "ocpu" in metric:
                    mem_sku = CPU_SKU_TO_MEM_SKU.get(sku, "—")
                    mem_row = price_table.get(mem_sku, {})
                    lines.append(
                        f"{label} | {sku} | {self._format_catalog_price(row.get('unit_price'))} "
                        f"| {mem_sku} | {self._format_catalog_price(mem_row.get('unit_price'))}"
                    )
                    break
        return "\n".join(lines)

    @staticmethod
    def _format_catalog_price(value: Any) -> str:
        return f"${float(value):.4f}" if value not in (None, "") else "TBD"

    def chat(
        self,
        *,
        message: str,
        conversation: list[dict[str, str]] | None = None,
        trace_id: str,
        model_id: str,
        text_runner: Callable[[str, str], str] | None = None,
    ) -> dict[str, Any]:
        started = time.perf_counter()
        with self._lock:
            snap = self._cache

        if not snap:
            return {
                "type": "normal",
                "reply": "BOM data is not ready. Run /api/bom/refresh-data, then retry.",
                "trace_id": trace_id,
                "trace": {
                    "model_id": model_id,
                    "type": "normal",
                    "repair_attempts": 0,
                    "cache_ready": False,
                    "cache_source": "none",
                    "latency_ms": int((time.perf_counter() - started) * 1000),
                },
            }

        intent = self._classify_intent(message)
        shape_catalog = ""

        if intent == "normal":
            reply = (
                "I can build a BOM, validate SKUs/prices, and export XLSX. "
                "Ask for a costed BOM with workload sizing details (OCPU, memory, storage, LB)."
            )
            result_type = "normal"
            payload = None
        elif intent == "question":
            reply = (
                "Before finalizing, please confirm: target region, non-GPU vs GPU compute, "
                "and expected OCPU/memory/storage quantities."
            )
            result_type = "question"
            payload = None
        else:
            raw_payload = self._draft_bom_payload(message, snap.pricing_table)
            shape_catalog = str(raw_payload.pop("_shape_catalog", "") or "")
            repaired_payload, attempts, errors = self._repair_until_valid(raw_payload, snap.pricing_table)
            if errors:
                result_type = "question"
                reply = (
                    "I could not finalize a valid BOM after 3 repair attempts. "
                    "Please provide exact SKU-level sizing inputs."
                )
                payload = None
            else:
                result_type = "final"
                payload = self._normalize_payload(repaired_payload)
                reply = "Final BOM prepared. Review line items, then export JSON or XLSX."

        elapsed_ms = int((time.perf_counter() - started) * 1000)
        trace = {
            "model_id": model_id,
            "type": result_type,
            "repair_attempts": attempts if intent == "final" else 0,
            "cache_ready": True,
            "cache_source": snap.source,
            "latency_ms": elapsed_ms,
        }
        if intent == "final" and shape_catalog:
            trace["shape_catalog"] = shape_catalog
        response: dict[str, Any] = {
            "type": result_type,
            "reply": reply,
            "trace_id": trace_id,
            "trace": trace,
        }
        if payload is not None:
            response["json_bom"] = json.dumps(payload, ensure_ascii=False, indent=2)
            response["bom_payload"] = payload
            response["score"] = 1.0
        return response

    def generate_from_inputs(
        self,
        *,
        inputs: dict[str, Any],
        trace_id: str,
        model_id: str,
    ) -> dict[str, Any]:
        started = time.perf_counter()
        with self._lock:
            snap = self._cache

        if not snap:
            return {
                "type": "normal",
                "reply": "BOM data is not ready. Run /api/bom/refresh-data, then retry.",
                "trace_id": trace_id,
                "trace": {
                    "model_id": model_id,
                    "type": "normal",
                    "repair_attempts": 0,
                    "cache_ready": False,
                    "cache_source": "none",
                    "latency_ms": int((time.perf_counter() - started) * 1000),
                    "bom_request_mode": "structured_inputs",
                },
            }

        normalized, blockers = self._normalize_bom_inputs(inputs)
        if blockers:
            return {
                "type": "question",
                "reply": (
                    "BOM clarification required. I found structured BOM sizing context, "
                    "but could not normalize it safely: " + "; ".join(blockers)
                ),
                "trace_id": trace_id,
                "structured_inputs": inputs,
                "normalized_inputs": normalized,
                "normalization_blockers": blockers,
                "trace": {
                    "model_id": model_id,
                    "type": "question",
                    "repair_attempts": 0,
                    "cache_ready": True,
                    "cache_source": snap.source,
                    "latency_ms": int((time.perf_counter() - started) * 1000),
                    "bom_request_mode": "structured_inputs",
                    "normalization_blockers": blockers,
                },
            }

        raw_payload = self._draft_bom_payload_from_inputs(normalized, snap.pricing_table)
        repaired_payload, attempts, errors = self._repair_until_valid(raw_payload, snap.pricing_table)
        if errors:
            return {
                "type": "question",
                "reply": (
                    "I could not finalize a valid BOM from the structured inputs after 3 repair attempts. "
                    "Please provide exact SKU-level sizing inputs."
                ),
                "trace_id": trace_id,
                "structured_inputs": inputs,
                "normalized_inputs": normalized,
                "normalization_blockers": errors,
                "trace": {
                    "model_id": model_id,
                    "type": "question",
                    "repair_attempts": attempts,
                    "cache_ready": True,
                    "cache_source": snap.source,
                    "latency_ms": int((time.perf_counter() - started) * 1000),
                    "bom_request_mode": "structured_inputs",
                    "normalization_blockers": errors,
                },
            }

        payload = self._normalize_payload(repaired_payload)
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        return {
            "type": "final",
            "reply": "Final BOM prepared from structured inputs. Review line items, then export JSON or XLSX.",
            "trace_id": trace_id,
            "structured_inputs": inputs,
            "normalized_inputs": normalized,
            "json_bom": json.dumps(payload, ensure_ascii=False, indent=2),
            "bom_payload": payload,
            "score": 1.0,
            "trace": {
                "model_id": model_id,
                "type": "final",
                "repair_attempts": attempts,
                "cache_ready": True,
                "cache_source": snap.source,
                "latency_ms": elapsed_ms,
                "bom_request_mode": "structured_inputs",
            },
        }

    def _normalize_bom_inputs(self, inputs: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
        src = inputs if isinstance(inputs, dict) else {}
        blockers: list[str] = []
        compute = src.get("compute", {}) if isinstance(src.get("compute"), dict) else {}
        memory = src.get("memory", {}) if isinstance(src.get("memory"), dict) else {}
        storage = src.get("storage", {}) if isinstance(src.get("storage"), dict) else {}
        database = src.get("database", {}) if isinstance(src.get("database"), dict) else {}
        connectivity = src.get("connectivity", {}) if isinstance(src.get("connectivity"), dict) else {}
        dr = src.get("dr", {}) if isinstance(src.get("dr"), dict) else {}
        target_service_ids = [
            str(item).strip()
            for item in src.get("target_service_ids", []) or []
            if str(item).strip()
        ]

        ocpu = self._normalize_plain_number(compute.get("ocpu"))
        memory_gb = self._normalize_capacity_gb(memory.get("gb"), default_unit="gb")
        block_tb = self._normalize_capacity_tb(storage.get("block_tb"), default_unit="tb")
        object_tb = self._normalize_capacity_tb(storage.get("object_tb"), default_unit="tb")
        file_tb = self._normalize_capacity_tb(storage.get("file_tb"), default_unit="tb")
        tiers = [
            dict(item)
            for item in compute.get("tiers", []) or []
            if isinstance(item, dict)
        ]

        requirements = []
        if not target_service_ids or any(item.startswith("compute.") for item in target_service_ids):
            requirements.extend((
                ("compute.ocpu", compute.get("ocpu"), ocpu),
                ("memory.gb", memory.get("gb"), memory_gb),
            ))
        if not target_service_ids or "storage.block" in target_service_ids:
            requirements.append(("storage.block_tb", storage.get("block_tb"), block_tb))
        for label, value, normalized in requirements:
            if value not in (None, "", [], {}) and normalized is None:
                blockers.append(f"{label}={value!r}")

        required_normalized = []
        if not target_service_ids or any(item.startswith("compute.") for item in target_service_ids):
            required_normalized.extend((("compute.ocpu", ocpu), ("memory.gb", memory_gb)))
        if not target_service_ids or "storage.block" in target_service_ids:
            required_normalized.append(("storage.block_tb", block_tb))
        for label, normalized in required_normalized:
            if normalized is None or normalized <= 0:
                blockers.append(f"missing required {label}")

        internet_mbps = self._normalize_plain_number(connectivity.get("internet_mbps"))
        rto_hours = self._normalize_plain_number(dr.get("rto_hours"))

        normalized = {
            "region": str(src.get("region", "") or "").strip(),
            "architecture_option": str(src.get("architecture_option", "") or "").strip(),
            "compute": {
                "ocpu": ocpu,
                "gpu": bool(compute.get("gpu", False)),
                "instance_count": int(compute.get("instance_count") or 0),
                "tiers": tiers,
            },
            "memory": {"gb": memory_gb},
            "storage": {"block_tb": block_tb, "block_gb": (block_tb * 1024.0) if block_tb is not None else None},
            "connectivity": {
                "internet_mbps": internet_mbps,
                "mpls": bool(connectivity.get("mpls", False)),
                "sd_wan": bool(connectivity.get("sd_wan", False)),
            },
            "dr": {
                "rto_hours": rto_hours,
                "cross_region_restore": bool(dr.get("cross_region_restore", False)),
            },
            "workloads": [str(item).strip() for item in src.get("workloads", []) or [] if str(item).strip()],
            "os_mix": [str(item).strip() for item in src.get("os_mix", []) or [] if str(item).strip()],
            "target_services": [str(item).strip() for item in src.get("target_services", []) or [] if str(item).strip()],
            "target_service_ids": target_service_ids,
            "strict_scope": bool(src.get("strict_scope", False)),
            "scope": str(src.get("scope", "") or "").strip(),
            "assumptions": [str(item).strip() for item in src.get("assumptions", []) or [] if str(item).strip()],
            "workload_service_mapping": [
                dict(item)
                for item in src.get("workload_service_mapping", []) or []
                if isinstance(item, dict)
            ],
            "output_format": str(src.get("output_format", "") or "xlsx").strip(),
            "database": {
                "service_id": str(database.get("service_id") or "").strip(),
                "display_name": str(database.get("display_name") or "").strip(),
                "ocpu": self._normalize_plain_number(database.get("ocpu")),
                "instance_count": int(database.get("instance_count") or 1),
                "ha": bool(database.get("ha", False)),
            },
        }
        normalized["storage"]["object_tb"] = object_tb
        normalized["storage"]["object_gb"] = (object_tb * 1024.0) if object_tb is not None else None
        normalized["storage"]["file_tb"] = file_tb
        normalized["storage"]["file_gb"] = (file_tb * 1024.0) if file_tb is not None else None
        return normalized, list(dict.fromkeys(blockers))

    @classmethod
    def _normalize_plain_number(cls, value: Any) -> float | None:
        if value in (None, "", [], {}):
            return None
        parsed = cls._parse_numeric(value)
        return parsed if parsed is not None and parsed > 0 else None

    @classmethod
    def _normalize_capacity_gb(cls, value: Any, *, default_unit: str) -> float | None:
        if value in (None, "", [], {}):
            return None
        if isinstance(value, (int, float)):
            numeric = float(value)
            if numeric <= 0:
                return None
            return numeric * 1024.0 if default_unit == "tb" else numeric
        text = str(value).replace(",", "").strip().lower()
        match = re.search(r"(\d+(?:\.\d+)?)\s*(tb|gb)?\b", text)
        if not match:
            return None
        numeric = float(match.group(1))
        unit = match.group(2) or default_unit
        gb = numeric * 1024.0 if unit == "tb" else numeric
        return gb if gb > 0 else None

    @classmethod
    def _normalize_capacity_tb(cls, value: Any, *, default_unit: str) -> float | None:
        gb = cls._normalize_capacity_gb(value, default_unit=default_unit)
        if gb is None:
            return None
        return gb / 1024.0

    def _draft_bom_payload_from_inputs(
        self,
        normalized: dict[str, Any],
        price_table: dict[str, dict[str, Any]],
    ) -> dict[str, Any]:
        if normalized.get("strict_scope") and normalized.get("target_service_ids"):
            return self._draft_strict_poc_payload(normalized, price_table)
        compute = normalized.get("compute", {}) if isinstance(normalized.get("compute"), dict) else {}
        memory = normalized.get("memory", {}) if isinstance(normalized.get("memory"), dict) else {}
        storage = normalized.get("storage", {}) if isinstance(normalized.get("storage"), dict) else {}
        connectivity = normalized.get("connectivity", {}) if isinstance(normalized.get("connectivity"), dict) else {}
        dr = normalized.get("dr", {}) if isinstance(normalized.get("dr"), dict) else {}
        target_services = [str(item).strip() for item in normalized.get("target_services", []) or [] if str(item).strip()]
        workloads = [str(item).strip() for item in normalized.get("workloads", []) or [] if str(item).strip()]
        target_text = " ".join([str(normalized.get("architecture_option", "") or ""), " ".join(target_services), " ".join(workloads)]).lower()
        is_native = "native" in target_text

        is_gpu = bool(compute.get("gpu"))
        ocpu = float(compute.get("ocpu") or 0.0)
        mem_gb = float(memory.get("gb") or 0.0)
        block_gb = float(storage.get("block_gb") or 0.0)
        object_gb = float(storage.get("object_gb") or 0.0)
        cpu_sku = (
            "B97384" if "e5.flex" in target_text or "standard.e5" in target_text
            else "B111129"
        )
        mem_sku = CPU_SKU_TO_MEM_SKU[cpu_sku]

        line_items = []
        cpu_line = self._build_line(cpu_sku, ocpu, price_table, "compute", "Structured input: compute.ocpu")
        if is_native:
            shape_label = "VM.Standard.E5.Flex" if cpu_sku == "B97384" else "VM.Standard.E6.Flex"
            cpu_line["description"] = f"{shape_label} compute VMs - OCPU"
            cpu_line["notes"] = f"OCI Native Services target: {shape_label} compute VMs for application, SQL Server, and lift/shift workloads"
        line_items.append(cpu_line)
        if not is_gpu:
            mem_line = self._build_line(mem_sku, mem_gb, price_table, "compute", "Structured input: memory.gb")
            if is_native:
                mem_line["description"] = f"{shape_label} compute VMs - Memory"
                mem_line["notes"] = f"OCI Native Services target: {shape_label} memory for compute VMs"
            line_items.append(mem_line)
        line_items.append(
            self._build_line("B91961", block_gb, price_table, "storage", "Structured input: storage.block_tb")
        )
        if object_gb > 0:
            line_items.append(
                self._build_line(
                    "B91628",
                    object_gb,
                    price_table,
                    "storage",
                    "Structured input: storage.object_tb",
                )
            )
        line_items.append(
            self._build_line(
                "B91962",
                self._block_volume_performance_units(block_gb),
                price_table,
                "storage",
                "Block Volume performance units; Balanced 10 VPU/GB from structured storage input",
            )
        )
        if is_native:
            line_items.extend(
                self._native_service_lines_from_inputs(
                    normalized=normalized,
                    price_table=price_table,
                    block_gb=block_gb,
                    object_gb=object_gb,
                )
            )

        assumptions = [
            "Pricing is estimate-only and non-binding.",
            "Monthly costs assume steady-state usage.",
            "Structured BOM inputs were used as the source of truth for compute, memory, and block storage sizing.",
        ]
        if normalized.get("region"):
            assumptions.append(f"Requested OCI region: {normalized['region']}.")
        if normalized.get("architecture_option"):
            assumptions.append(f"Architecture option: {normalized['architecture_option']}.")
        if is_native:
            assumptions.append("OCI Native Services selected; do not model OCVS, vCenter, NSX, or ESXi as target components.")
        if target_services:
            assumptions.append("Native target services: " + ", ".join(target_services) + ".")
        if normalized.get("workload_service_mapping"):
            formatted = [
                f"{item.get('workload')} -> {item.get('target_service')}"
                for item in normalized.get("workload_service_mapping", [])[:8]
                if isinstance(item, dict) and item.get("workload") and item.get("target_service")
            ]
            if formatted:
                assumptions.append("Workload-to-service mapping: " + "; ".join(formatted) + ".")
        if connectivity.get("internet_mbps"):
            assumptions.append(f"Internet connectivity target: {connectivity['internet_mbps']:g} Mbps.")
        if connectivity.get("mpls"):
            assumptions.append("MPLS connectivity requirement captured for architecture review.")
        if connectivity.get("sd_wan"):
            assumptions.append("SD-WAN connectivity requirement captured for architecture review.")
        if dr.get("rto_hours"):
            assumptions.append(f"DR RTO/restore target: {dr['rto_hours']:g} hours.")
        if dr.get("cross_region_restore"):
            assumptions.append("Cross-region restore requirement captured for architecture review.")

        resolved_inputs = self._resolved_inputs_from_structured_bom(normalized)
        payload = {
            "currency": "USD",
            "region": normalized.get("region", ""),
            "architecture_option": normalized.get("architecture_option", ""),
            "line_items": line_items,
            "assumptions": assumptions,
            "resolved_inputs": resolved_inputs,
            "workloads": list(normalized.get("workloads", []) or []),
            "os_mix": list(normalized.get("os_mix", []) or []),
            "structured_inputs": normalized,
        }
        return payload

    def _draft_strict_poc_payload(
        self,
        normalized: dict[str, Any],
        price_table: dict[str, dict[str, Any]],
    ) -> dict[str, Any]:
        service_ids = set(normalized.get("target_service_ids", []) or [])
        compute = normalized.get("compute", {}) if isinstance(normalized.get("compute"), dict) else {}
        memory = normalized.get("memory", {}) if isinstance(normalized.get("memory"), dict) else {}
        storage = normalized.get("storage", {}) if isinstance(normalized.get("storage"), dict) else {}
        rows: list[dict[str, Any]] = []

        compute_id = next((item for item in service_ids if item.startswith("compute.")), "")
        if compute_id:
            cpu_sku = "B97384" if compute_id == "compute.vm.standard.e5.flex" else "B111129"
            mem_sku = CPU_SKU_TO_MEM_SKU[cpu_sku]
            tiers = [item for item in compute.get("tiers", []) or [] if isinstance(item, dict)]
            if tiers:
                for tier in tiers:
                    tier_name = str(tier.get("name") or "compute tier").strip()
                    count = int(tier.get("instance_count") or 1)
                    for sku, quantity, resource in (
                        (cpu_sku, float(tier.get("ocpu") or 0), "OCPU"),
                        (mem_sku, float(tier.get("memory_gb") or 0), "Memory"),
                    ):
                        row = self._build_line(
                            sku,
                            quantity,
                            price_table,
                            "compute",
                            f"Explicit {tier_name}: {count} servers; {resource} sized per tier",
                        )
                        row["description"] = f"{consistency_contract.display_name(compute_id)} {tier_name} tier - {resource}"
                        row["canonical_service_id"] = compute_id
                        row["instance_count"] = count
                        rows.append(row)
            else:
                for sku, quantity, notes in (
                    (cpu_sku, float(compute.get("ocpu") or 4.0), "Assumed minimum POC compute OCPU"),
                    (mem_sku, float(memory.get("gb") or 32.0), "Assumed minimum POC compute memory"),
                ):
                    row = self._build_line(sku, quantity, price_table, "compute", notes)
                    row["canonical_service_id"] = compute_id
                    row["instance_count"] = int(compute.get("instance_count") or 1)
                    rows.append(row)
        if "storage.block" in service_ids:
            block_gb = float(storage.get("block_gb") or 500.0)
            for sku, quantity, notes in (
                ("B91961", block_gb, "Assumed minimum POC Balanced Block Volume"),
                ("B91962", self._block_volume_performance_units(block_gb), "Balanced 10 VPU/GB"),
            ):
                row = self._build_line(sku, quantity, price_table, "storage", notes)
                row["canonical_service_id"] = "storage.block"
                rows.append(row)
        if "storage.object" in service_ids:
            row = self._build_line(
                "B91628", float(storage.get("object_gb") or 1024.0), price_table,
                "storage", "Assumed minimum POC Object Storage",
            )
            row["canonical_service_id"] = "storage.object"
            rows.append(row)
        if "storage.file" in service_ids:
            row = self._build_line(
                "BFILE01", float(storage.get("file_gb") or 1024.0), price_table,
                "storage", "Explicit or assumed File Storage capacity",
            )
            row["canonical_service_id"] = "storage.file"
            rows.append(row)
        if "network.load_balancer.flexible" in service_ids:
            row = self._build_line("B93030", 1.0, price_table, "network", "One selected Flexible Load Balancer")
            row["canonical_service_id"] = "network.load_balancer.flexible"
            rows.append(row)
        if "security.waf" in service_ids:
            row = self._build_line("BWAF01", 1.0, price_table, "security", "Selected OCI WAF; authoritative price required")
            row["canonical_service_id"] = "security.waf"
            rows.append(row)
        database = normalized.get("database", {}) if isinstance(normalized.get("database"), dict) else {}
        database_id = str(database.get("service_id") or "")
        selected_database_id = next((item for item in service_ids if item.startswith("database.")), "")
        if selected_database_id:
            database_ocpu = float(database.get("ocpu") or 2.0)
            row = self._build_line(
                "B99060", database_ocpu, price_table, "database",
                f"{'Explicit HA' if database.get('ha') else 'Assumed'} {consistency_contract.display_name(selected_database_id)} sizing",
            )
            row["description"] = f"{consistency_contract.display_name(selected_database_id)} - OCPU"
            row["canonical_service_id"] = selected_database_id
            row["instance_count"] = 1
            rows.append(row)

        return {
            "currency": "USD",
            "region": str(normalized.get("region") or ""),
            "architecture_option": str(normalized.get("architecture_option") or ""),
            "line_items": rows,
            "assumptions": list(normalized.get("assumptions", []) or []),
            "structured_inputs": normalized,
        }

    def _native_service_lines_from_inputs(
        self,
        *,
        normalized: dict[str, Any],
        price_table: dict[str, dict[str, Any]],
        block_gb: float,
        object_gb: float,
    ) -> list[dict[str, Any]]:
        connectivity = normalized.get("connectivity", {}) if isinstance(normalized.get("connectivity"), dict) else {}
        dr = normalized.get("dr", {}) if isinstance(normalized.get("dr"), dict) else {}
        workloads = [str(item).strip() for item in normalized.get("workloads", []) or [] if str(item).strip()]
        target_services = [str(item).strip() for item in normalized.get("target_services", []) or [] if str(item).strip()]
        text = " ".join(
            [str(normalized.get("architecture_option", "") or ""), " ".join(workloads), " ".join(target_services)]
        ).lower()
        rows: list[dict[str, Any]] = []

        def _has(*markers: str) -> bool:
            return any(marker in text for marker in markers)

        if _has("autonomous database", "oracle database", "oracle databases", "oracle db", "database", "data tier", "db", "adb", "atp", "adw"):
            rows.append(
                self._build_line(
                    "B99060",
                    max(2.0, float((normalized.get("compute", {}) or {}).get("ocpu") or 0.0) * 0.25),
                    price_table,
                    "database",
                    "OCI Native Services target: Autonomous Database for Oracle database workloads",
                )
            )
        if _has("file storage", "file server", "file servers", "file share", "file shares", "nfs", "smb"):
            rows.append(
                self._build_line(
                    "BFILE01",
                    max(1024.0, block_gb * 0.25),
                    price_table,
                    "storage",
                    "OCI Native Services target: File Storage for file shares and file server workloads",
                )
            )
        if _has("load balancer", "waf", "web", "http", "https", "ingress", "public"):
            rows.append(
                self._build_line(
                    "B93030",
                    1.0,
                    price_table,
                    "network",
                    "OCI Native Services target: Load Balancer for application ingress",
                )
            )
            rows.append(
                self._build_line(
                    "BWAF01",
                    1.0,
                    price_table,
                    "network",
                    "OCI Native Services target: WAF policy for protected ingress",
                )
            )
        if object_gb <= 0 and (_has("object storage", "backup", "archive") or dr.get("rto_hours") or dr.get("cross_region_restore")):
            rows.append(
                self._build_line(
                    "B91628",
                    max(1024.0, block_gb * 0.2),
                    price_table,
                    "storage",
                    "OCI Native Services target: Object Storage backups/archive and recovery copies",
                )
            )
        if connectivity.get("mpls") or connectivity.get("sd_wan") or _has("drg", "fastconnect", "mpls", "sd-wan", "sd wan", "vpn"):
            rows.append(
                self._build_line(
                    "B88325",
                    1.0,
                    price_table,
                    "network",
                    "OCI Native Services target: DRG/FastConnect/MPLS private connectivity",
                )
            )
        return rows

    @staticmethod
    def _resolved_inputs_from_structured_bom(normalized: dict[str, Any]) -> list[dict[str, str]]:
        compute = normalized.get("compute", {}) if isinstance(normalized.get("compute"), dict) else {}
        memory = normalized.get("memory", {}) if isinstance(normalized.get("memory"), dict) else {}
        storage = normalized.get("storage", {}) if isinstance(normalized.get("storage"), dict) else {}
        connectivity = normalized.get("connectivity", {}) if isinstance(normalized.get("connectivity"), dict) else {}
        dr = normalized.get("dr", {}) if isinstance(normalized.get("dr"), dict) else {}
        rows: list[tuple[str, Any, str]] = [
            ("bom.region", normalized.get("region"), "{value}"),
            ("bom.architecture_option", normalized.get("architecture_option"), "{value}"),
            ("bom.compute.ocpu", compute.get("ocpu"), "{value:g} OCPU equivalent"),
            ("bom.compute.gpu", "GPU compute" if compute.get("gpu") else "non-GPU compute", "{value}"),
            ("bom.compute.memory", memory.get("gb"), "{value:g} GB RAM"),
            ("bom.storage.block", storage.get("block_tb"), "{value:g} TB block storage"),
            ("bom.connectivity.internet", connectivity.get("internet_mbps"), "{value:g} Mbps internet"),
            ("bom.connectivity.mpls", "MPLS" if connectivity.get("mpls") else "", "{value}"),
            ("bom.connectivity.sd_wan", "SD-WAN" if connectivity.get("sd_wan") else "", "{value}"),
            ("bom.dr.rto", dr.get("rto_hours"), "{value:g} hour RTO/restore target"),
            ("bom.dr.cross_region_restore", "cross-region restore" if dr.get("cross_region_restore") else "", "{value}"),
        ]
        resolved: list[dict[str, str]] = []
        for question_id, value, template in rows:
            if value in (None, "", [], {}):
                continue
            answer = template.format(value=float(value) if isinstance(value, (int, float)) else str(value))
            resolved.append({"question_id": question_id, "answer": answer, "source": "structured_bom_inputs"})
        if normalized.get("workloads"):
            resolved.append(
                {
                    "question_id": "bom.workloads",
                    "answer": ", ".join(str(item) for item in normalized.get("workloads", [])[:8]),
                    "source": "structured_bom_inputs",
                }
            )
        if normalized.get("os_mix"):
            resolved.append(
                {
                    "question_id": "bom.os_mix",
                    "answer": ", ".join(str(item) for item in normalized.get("os_mix", [])[:8]),
                    "source": "structured_bom_inputs",
                }
            )
        return resolved

    def _classify_intent(self, message: str) -> str:
        m = (message or "").strip().lower()
        if not m:
            return "normal"
        final_markers = (
            "final",
            "generate bom",
            "bill of materials",
            "cost estimate",
            "price bom",
            "export bom",
        )
        question_markers = ("how", "what", "?", "clarify", "need info")
        sizing_markers = ("ocpu", "gb", "tb", "gpu", "load balancer", "storage")

        if any(marker in m for marker in final_markers):
            return "final"
        if "bom" in m and any(marker in m for marker in sizing_markers):
            return "final"
        if "bom" in m and any(marker in m for marker in question_markers):
            return "question"
        if "bom" in m or "pricing" in m or "cost" in m:
            return "question"
        return "normal"

    @staticmethod
    def _user_request_text(message: str) -> str:
        """Return only the user-supplied portion of message, stripping any Archie Canonical Memory block."""
        if "[End Archie Canonical Memory]" in message:
            parts = message.split("[End Archie Canonical Memory]", 1)
            return parts[1].strip()
        return message

    def _draft_bom_payload(self, message: str, price_table: dict[str, dict[str, Any]]) -> dict[str, Any]:
        # Use only the user-supplied portion for shape/SKU hints so that old
        # E4 descriptions in the canonical memory block don't force E4 selection.
        user_text = self._user_request_text(message).lower()
        text = message.lower()
        is_gpu = "gpu" in text
        mentions_non_oci = self._mentions_non_oci_provider(text)
        # Canonical memory and reviewer notes may contain older tables or
        # defaults.  Explicit request sizing must win, so derive table signals
        # from the user portion only.
        table_signals = self._extract_table_signals(user_text)

        # Detect instance count ("2 servers", "3 nodes", etc.) from user text only.
        _server_count = self._extract_server_count(user_text)

        ocpu = float(table_signals.get("ocpu") or 0.0)
        if ocpu <= 0:
            ocpu = self._extract_number(r"(\d+(?:\.\d+)?)\s*ocpu", user_text, default=4.0)

        mem_gb = float(table_signals.get("mem_gb") or 0.0)
        if mem_gb <= 0:
            mem_gb = self._extract_number(
                r"(\d+(?:\.\d+)?)\s*gb\s*(?:memory|ram)?",
                user_text,
                default=ocpu * 16,
            )

        block_gb = float(table_signals.get("block_gb") or 0.0)
        if block_gb <= 0:
            block_match = re.search(
                r"(\d+(?:\.\d+)?)\s*(gb|tb)\s*(?:balanced\s+|higher\s+performance\s+)?"
                r"(?:block\s+(?:volume|storage)|boot\s+volume)",
                user_text,
            )
            if not block_match:
                block_match = re.search(
                    r"(?:block\s+(?:volume|storage)|boot\s+volume)[^\d]{0,24}"
                    r"(\d+(?:\.\d+)?)\s*(gb|tb)",
                    user_text,
                )
            if block_match:
                block_gb = float(block_match.group(1)) * (
                    1024.0 if block_match.group(2).lower() == "tb" else 1.0
                )
            else:
                block_tb = self._extract_number(
                    r"(\d+(?:\.\d+)?)\s*tb\s*(?:block|storage)",
                    user_text,
                    default=1.0,
                )
                block_gb = block_tb * 1024.0

        object_storage_gb = float(table_signals.get("object_storage_gb") or 0.0)
        if object_storage_gb <= 0:
            object_match = re.search(
                r"(\d+(?:\.\d+)?)\s*(gb|tb)\s*(?:standard\s+)?object\s+storage",
                user_text,
            )
            if not object_match:
                object_match = re.search(
                    r"object\s+storage[^\d]{0,24}(\d+(?:\.\d+)?)\s*(gb|tb)",
                    user_text,
                )
            if object_match:
                object_storage_gb = float(object_match.group(1)) * (
                    1024.0 if object_match.group(2).lower() == "tb" else 1.0
                )
        load_balancer_qty = float(table_signals.get("load_balancer_qty") or 0.0)
        ocpu_notes = str(table_signals.get("ocpu_notes") or "Primary compute OCPU")
        mem_notes = str(table_signals.get("mem_notes") or "Primary compute memory")
        block_notes = str(table_signals.get("block_notes") or "Block storage capacity")

        # Detect "N servers/instances/nodes" and multiply per-server sizing
        _server_count = self._extract_server_count(user_text)
        if _server_count > 1:
            _per_ocpu = ocpu
            _per_mem  = mem_gb
            ocpu   = _per_ocpu * _server_count
            mem_gb = _per_mem  * _server_count
            ocpu_notes = (
                f"Compute OCPU — {_server_count} servers × {int(_per_ocpu)} OCPU"
            )
            mem_notes  = (
                f"Compute memory — {_server_count} servers × {int(_per_mem)} GB"
            )

        shape_hint = str(table_signals.get("cpu_family") or "").lower()
        if not shape_hint:
            if "ampere" in user_text or "a1" in user_text:
                shape_hint = "a1"
            elif "e6" in user_text:
                shape_hint = "e6"

        # Determine CPU SKU from user request only; default to E6 for unspecified compute.
        if shape_hint == "a1" or "ampere" in user_text:
            cpu_sku = "B93297"
        elif "e6" in user_text or shape_hint == "e6":
            cpu_sku = "B111129"
        elif "e4" in user_text or shape_hint == "e4":
            cpu_sku = "B93113"
        elif "e5" in user_text or shape_hint == "e5":
            cpu_sku = "B97384"
        elif "x9" in user_text or "intel" in user_text:
            x9_sku = "B94176"
            cpu_sku = x9_sku
        else:
            cpu_sku = "B111129"   # E6.Flex — OCI default general-purpose VM
        mem_sku = CPU_SKU_TO_MEM_SKU[cpu_sku]
        shape_catalog = self._build_shape_catalog(price_table)

        line_items: list[dict[str, Any]] = []
        cpu_line = self._build_line(cpu_sku, ocpu, price_table, "compute", ocpu_notes)
        cpu_line["instance_count"] = _server_count
        line_items.append(cpu_line)

        # non-GPU compute must be split into OCPU + memory rows
        if not is_gpu:
            mem_line = self._build_line(mem_sku, mem_gb, price_table, "compute", mem_notes)
            mem_line["instance_count"] = _server_count
            line_items.append(mem_line)

        line_items.append(self._build_line("B91961", block_gb, price_table, "storage", block_notes))
        line_items.append(
            self._build_line(
                "B91962",
                self._block_volume_performance_units(block_gb),
                price_table,
                "storage",
                "Block Volume performance units; default Balanced 10 VPU/GB",
            )
        )

        if load_balancer_qty > 0 or "load balancer" in text or "lb" in text or "ingress" in text:
            line_items.append(
                self._build_line(
                    "B93030",
                    load_balancer_qty or 1.0,
                    price_table,
                    "network",
                    "Flexible load balancer",
                )
            )

        if "waf" in text or "web application firewall" in text:
            line_items.append(
                self._build_line(
                    "BWAF01",
                    1.0,
                    price_table,
                    "network",
                    "Web Application Firewall policy",
                )
            )

        if "database" in text or "data tier" in text or re.search(r"\bdb\b", text):
            line_items.append(
                self._build_line(
                    "B99060",
                    max(2.0, ocpu * 0.25),
                    price_table,
                    "database",
                    "Database layer",
                )
            )

        if object_storage_gb > 0 or "object storage" in text:
            line_items.append(
                self._build_line(
                    "B91628",
                    object_storage_gb or max(100.0, block_gb * 0.2),
                    price_table,
                    "storage",
                    "Object storage",
                )
            )

        assumptions = [
            "Pricing is estimate-only and non-binding.",
            "Monthly costs assume steady-state usage.",
        ]
        if mentions_non_oci:
            assumptions.append(
                "OCI-only BOM enforced: non-OCI provider references in the request were normalized to OCI equivalents."
            )

        return {
            "currency": "USD",
            "region": self._extract_region(user_text),
            "line_items": line_items,
            "assumptions": assumptions,
            "_shape_catalog": shape_catalog,
        }

    @staticmethod
    def _extract_number(pattern: str, text: str, default: float) -> float:
        match = re.search(pattern, text)
        if not match:
            return default
        try:
            return float(match.group(1))
        except Exception:
            return default

    @staticmethod
    def _extract_server_count(text: str) -> int:
        descriptor = r"(?:(?:vm\.[a-z0-9_.-]+|compute|application|app|web|worker|database)\s+){0,4}"
        match = re.search(
            rf"\b(\d+)\s+(?:private\s+|public\s+)?{descriptor}(?:server|instance|vm|node|host)s?\b",
            text,
        )
        if match:
            return max(1, int(match.group(1)))
        word_values = {
            "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
            "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
        }
        word_match = re.search(
            r"\b(one|two|three|four|five|six|seven|eight|nine|ten)\s+"
            rf"(?:private\s+|public\s+)?{descriptor}(?:server|instance|vm|node|host)s?\b",
            text,
        )
        return word_values.get(word_match.group(1), 1) if word_match else 1

    @staticmethod
    def _extract_region(text: str) -> str:
        match = re.search(r"\b[a-z]{2,}-[a-z]+-\d+\b", text, re.IGNORECASE)
        return match.group(0) if match else ""

    @staticmethod
    def _build_line(
        sku: str,
        quantity: float,
        price_table: dict[str, dict[str, Any]],
        category: str,
        notes: str,
    ) -> dict[str, Any]:
        row = price_table.get(sku, {"description": sku, "unit_price": None, "source": "unpriced_catalog"})
        raw_price = row.get("unit_price")
        unit_price = float(raw_price) if raw_price not in (None, "") else None
        qty = float(quantity)
        metric = str(row.get("metric") or "")
        monthly_multiplier = BomService._monthly_multiplier_for_metric(metric)
        return {
            "sku": sku,
            "description": str(row.get("description") or sku),
            "category": category,
            "metric": metric,
            "quantity": qty,
            "unit_price": unit_price,
            "monthly_multiplier": monthly_multiplier,
            "extended_price": round(qty * unit_price * monthly_multiplier, 4) if unit_price is not None else None,
            "price_source": str(row.get("source") or "unpriced_catalog"),
            "pricing_status": "priced" if unit_price is not None else "unpriced",
            "notes": notes,
        }

    @staticmethod
    def _block_volume_performance_units(block_gb: float, vpu_per_gb: float = 10.0) -> float:
        return max(0.0, float(block_gb) * vpu_per_gb)

    @staticmethod
    def _monthly_multiplier_for_metric(metric: str) -> int:
        lowered = metric.lower()
        return 730 if "hour" in lowered else 1

    @staticmethod
    def _mentions_non_oci_provider(text: str) -> bool:
        normalized = f" {text.strip().lower()} "
        return any(pattern in normalized for pattern in NON_OCI_PROVIDER_PATTERNS)

    def _extract_table_signals(self, message: str) -> dict[str, float | str]:
        signals: dict[str, float | str] = {
            "ocpu": 0.0,
            "mem_gb": 0.0,
            "block_gb": 0.0,
            "object_storage_gb": 0.0,
            "load_balancer_qty": 0.0,
            "cpu_family": "",
            "ocpu_notes": "",
            "mem_notes": "",
            "block_notes": "",
        }

        for row in self._parse_markdown_table_rows(message):
            compact_resource_row = len(row) == 3
            category = row[0].lower() if len(row) > 0 else ""
            component = "" if compact_resource_row else (row[1].lower() if len(row) > 1 else "")
            details = " | ".join(row[1:]).lower() if compact_resource_row else (row[2].lower() if len(row) > 2 else "")
            qty_cell = "" if compact_resource_row else (row[3] if len(row) > 3 else "")
            qty = self._extract_table_quantity(qty_cell)
            row_text = " | ".join(row).lower()

            if self._is_compute_row(category, component, row_text) or self._is_memory_row(row_text):
                row_ocpu = self._extract_ocpu(row_text)
                row_mem_gb = self._extract_memory_gb(row_text)
                row_block_gb = self._extract_keyword_capacity_gb(details, keywords=("block", "volume", "vol", "boot"))
                if row_ocpu > 0:
                    signals["ocpu"] = float(signals["ocpu"]) + (row_ocpu * qty)
                    signals["ocpu_notes"] = self._merge_note(
                        str(signals.get("ocpu_notes") or ""),
                        self._resource_note("Primary compute OCPU", row_text),
                    )
                if row_mem_gb > 0:
                    signals["mem_gb"] = float(signals["mem_gb"]) + (row_mem_gb * qty)
                    signals["mem_notes"] = self._merge_note(
                        str(signals.get("mem_notes") or ""),
                        self._memory_note(row_text, row_mem_gb),
                    )
                signals["block_gb"] = float(signals["block_gb"]) + (row_block_gb * qty)
                if "ampere" in row_text or "a1" in row_text or "arm" in row_text:
                    signals["cpu_family"] = "a1"
                elif "e6" in row_text and not signals["cpu_family"]:
                    signals["cpu_family"] = "e6"
                continue

            if "load balancer" in row_text:
                signals["load_balancer_qty"] = max(float(signals["load_balancer_qty"]), qty)
                continue

            if "object storage" in row_text:
                object_storage_gb = self._extract_first_non_egress_capacity_gb(details)
                if object_storage_gb > 0:
                    signals["object_storage_gb"] = max(float(signals["object_storage_gb"]), object_storage_gb * qty)
                continue

            if self._is_block_storage_row(row_text):
                block_gb = self._extract_keyword_capacity_gb(
                    row_text,
                    keywords=("block", "volume", "vol", "boot", "storage", "vsan", "hci", "capacity"),
                )
                if block_gb <= 0:
                    block_gb = self._extract_first_non_egress_capacity_gb(row_text)
                if block_gb > 0:
                    signals["block_gb"] = float(signals["block_gb"]) + (block_gb * qty)
                    signals["block_notes"] = self._merge_note(
                        str(signals.get("block_notes") or ""),
                        self._resource_note("Block storage capacity", row_text),
                    )

        return signals

    @staticmethod
    def _parse_markdown_table_rows(message: str) -> list[list[str]]:
        rows: list[list[str]] = []
        for raw_line in (message or "").splitlines():
            line = raw_line.strip()
            if not line.startswith("|") or line.count("|") < 3:
                continue
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if not cells:
                continue
            if all(re.fullmatch(r"[:\-\s]+", cell or "") for cell in cells):
                continue
            if cells[0].lower() in {"category", "resource"}:
                continue
            rows.append(cells)
        return rows

    @staticmethod
    def _extract_table_quantity(cell: str) -> float:
        match = re.search(r"(\d+(?:\.\d+)?)", cell or "")
        if not match:
            return 1.0
        try:
            value = float(match.group(1))
        except Exception:
            return 1.0
        return value if value > 0 else 1.0

    @staticmethod
    def _is_compute_row(category: str, component: str, row_text: str) -> bool:
        compute_markers = ("compute", "app server", "instance pool", "ampere", "a1 flex", "cpu", "ocpu")
        return any(marker in category for marker in compute_markers) or any(
            marker in component or marker in row_text for marker in compute_markers
        )

    @staticmethod
    def _is_memory_row(row_text: str) -> bool:
        return any(marker in row_text for marker in ("ram", "memory"))

    @staticmethod
    def _is_block_storage_row(row_text: str) -> bool:
        if "object storage" in row_text:
            return False
        storage_markers = ("block", "storage", "vsan", "hci", "capacity")
        return any(marker in row_text for marker in storage_markers) and bool(_CAPACITY_RE.search(row_text))

    @staticmethod
    def _capacity_to_gb(value: float, unit: str) -> float:
        return value * 1024.0 if unit.lower() == "tb" else value

    @staticmethod
    def _format_quantity(value: float, unit: str = "") -> str:
        rendered = str(int(value)) if float(value).is_integer() else str(value)
        return f"{rendered} {unit}".strip()

    @staticmethod
    def _merge_note(existing: str, addition: str) -> str:
        if not existing:
            return addition
        if addition in existing:
            return existing
        return f"{existing}; {addition}"

    @classmethod
    def _resource_note(cls, base: str, row_text: str) -> str:
        return f"{base}; source row: {row_text}"

    @classmethod
    def _memory_note(cls, row_text: str, target_gb: float) -> str:
        source_gb = cls._extract_source_memory_gb(row_text, target_gb)
        target_text = cls._format_quantity(target_gb, "GB")
        if source_gb > 0 and source_gb != target_gb:
            source_text = cls._format_quantity(source_gb, "GB")
            source_label = "source VxRail RAM" if "vxrail" in row_text else "source RAM"
            return f"Primary compute memory; {source_label} {source_text}; target OCI-equivalent RAM {target_text}"
        return cls._resource_note("Primary compute memory", row_text)

    @classmethod
    def _extract_ocpu(cls, text: str) -> float:
        patterns = (
            r"(?:oci[-\s]?equiv(?:alent)?|equiv(?:alent)?|target)[^\d]{0,24}~?\s*(\d+(?:\.\d+)?)\s*ocpu\b",
            r"~?\s*(\d+(?:\.\d+)?)\s*ocpu\b",
        )
        for pattern in patterns:
            value = cls._extract_number(pattern, text, default=0.0)
            if value > 0:
                return value
        return 0.0

    @classmethod
    def _extract_memory_gb(cls, text: str) -> float:
        target = cls._extract_equivalent_capacity_gb(text, context_markers=("ram", "memory"))
        if target > 0:
            return target

        for match in _CAPACITY_RE.finditer(text or ""):
            value = float(match.group(1))
            unit = match.group(2)
            start, end = match.span()
            before = text[max(0, start - 32):start].lower()
            after = text[end:min(len(text), end + 32)].lower()
            if any(marker in before or marker in after for marker in ("ram", "memory")):
                return cls._capacity_to_gb(value, unit)
        return 0.0

    @classmethod
    def _extract_source_memory_gb(cls, text: str, target_gb: float) -> float:
        for match in _CAPACITY_RE.finditer(text or ""):
            value = cls._capacity_to_gb(float(match.group(1)), match.group(2))
            if value == target_gb:
                continue
            start, end = match.span()
            before = text[max(0, start - 32):start].lower()
            after = text[end:min(len(text), end + 32)].lower()
            if any(marker in before or marker in after for marker in ("ram", "memory")):
                return value
        return 0.0

    @classmethod
    def _extract_equivalent_capacity_gb(cls, text: str, context_markers: tuple[str, ...]) -> float:
        equivalence = r"(?:oci[-\s]?equiv(?:alent)?|equiv(?:alent)?|target)"
        patterns = (
            rf"{equivalence}[^\d]{{0,32}}(\d+(?:\.\d+)?)\s*(tb|gb)\b",
            rf"(\d+(?:\.\d+)?)\s*(tb|gb)\b[^\n|]{{0,32}}{equivalence}",
        )
        for pattern in patterns:
            for match in re.finditer(pattern, text or "", flags=re.IGNORECASE):
                start, end = match.span()
                window = text[max(0, start - 48):min(len(text), end + 48)].lower()
                if not any(marker in window or marker in (text or "").lower() for marker in context_markers):
                    continue
                return cls._capacity_to_gb(float(match.group(1)), match.group(2))
        return 0.0

    @classmethod
    def _extract_first_non_egress_capacity_gb(cls, text: str) -> float:
        for match in _CAPACITY_RE.finditer(text or ""):
            value = float(match.group(1))
            unit = match.group(2)
            _, end = match.span()
            trailing = text[end:min(len(text), end + 8)].lower()
            if "egress" in trailing or "traffic" in trailing:
                continue
            return cls._capacity_to_gb(value, unit)
        return 0.0

    @classmethod
    def _extract_keyword_capacity_gb(cls, text: str, keywords: tuple[str, ...]) -> float:
        best_value = 0.0
        best_distance: int | None = None

        for match in _CAPACITY_RE.finditer(text or ""):
            value = float(match.group(1))
            unit = match.group(2)
            start, end = match.span()
            before = text[max(0, start - 24):start].lower()
            after = text[end:min(len(text), end + 24)].lower()
            trailing = after[:14]
            if "egress" in trailing or "traffic" in trailing:
                continue

            distances: list[int] = []
            for keyword in keywords:
                before_idx = before.rfind(keyword)
                if before_idx >= 0:
                    distances.append(len(before) - before_idx)
                after_idx = after.find(keyword)
                if after_idx >= 0:
                    distances.append(after_idx + 1)
            if not distances:
                continue

            distance = min(distances)
            if best_distance is None or distance < best_distance:
                best_distance = distance
                best_value = cls._capacity_to_gb(value, unit)

        return best_value

    def _repair_until_valid(
        self,
        payload: dict[str, Any],
        pricing_table: dict[str, dict[str, Any]],
    ) -> tuple[dict[str, Any], int, list[str]]:
        attempts = 0
        current = json.loads(json.dumps(payload))

        while attempts < 3:
            errors = self.validate_final_payload(current, pricing_table)
            if not errors:
                return current, attempts, []
            attempts += 1
            current = self.repair_payload(current, pricing_table, errors)

        final_errors = self.validate_final_payload(current, pricing_table)
        return current, attempts, final_errors

    def validate_final_payload(
        self,
        payload: dict[str, Any],
        pricing_table: dict[str, dict[str, Any]],
    ) -> list[str]:
        errors: list[str] = []
        line_items = payload.get("line_items")
        if not isinstance(line_items, list) or not line_items:
            return ["line_items must be a non-empty array"]

        seen_cpu = False
        seen_mem = False

        for idx, row in enumerate(line_items):
            if not isinstance(row, dict):
                errors.append(f"line_items[{idx}] must be object")
                continue
            sku = str(row.get("sku") or "").strip().upper()
            if sku not in pricing_table:
                errors.append(f"line_items[{idx}] unknown SKU: {sku}")
                continue
            ref = pricing_table[sku]
            raw_ref_price = ref.get("unit_price")
            ref_price = float(raw_ref_price) if raw_ref_price not in (None, "") else None
            raw_unit_price = row.get("unit_price")
            unit_price = float(raw_unit_price) if raw_unit_price not in (None, "") else None
            if ref_price is None:
                if unit_price is not None or str(row.get("pricing_status") or "") != "unpriced":
                    errors.append(f"line_items[{idx}] unverified price for SKU {sku} must remain unpriced")
            elif unit_price is None or unit_price <= 0:
                errors.append(f"line_items[{idx}] non-positive unit_price for SKU {sku}")
            elif abs(unit_price - ref_price) > 0.000001:
                errors.append(f"line_items[{idx}] unit_price for SKU {sku} does not match authoritative price")
            if ref_price is not None and ref_price < 0:
                errors.append(f"line_items[{idx}] negative authoritative unit_price for SKU {sku}")
            category = str(row.get("category") or "").lower()
            desc = str(row.get("description") or "").lower()
            notes = str(row.get("notes") or "").lower()
            if category == "compute" and "gpu" not in desc:
                if sku in CPU_SKU_TO_MEM_SKU:
                    seen_cpu = True
                if sku in CPU_SKU_TO_MEM_SKU.values():
                    seen_mem = True
            if self._mentions_non_oci_provider(desc) or self._mentions_non_oci_provider(notes):
                errors.append(f"line_items[{idx}] references non-OCI provider content")

        # non-GPU compute split rule
        if seen_cpu and not seen_mem:
            errors.append("non-GPU compute rows must include both OCPU and memory SKUs")

        return errors

    def repair_payload(
        self,
        payload: dict[str, Any],
        pricing_table: dict[str, dict[str, Any]],
        errors: list[str],
    ) -> dict[str, Any]:
        fixed = json.loads(json.dumps(payload))
        line_items: list[dict[str, Any]] = [
            row for row in fixed.get("line_items", []) if isinstance(row, dict)
        ]
        output: list[dict[str, Any]] = []

        for row in line_items:
            sku = str(row.get("sku") or "").strip().upper()
            if sku not in pricing_table:
                continue
            ref = pricing_table[sku]
            row["description"] = ref.get("description") or row.get("description") or sku
            row["category"] = ref.get("category") or row.get("category") or "service"
            row["metric"] = ref.get("metric") or row.get("metric") or ""
            try:
                quantity = float(row.get("quantity") or 0)
            except Exception:
                quantity = 0.0
            if quantity <= 0:
                quantity = 1.0
            row["quantity"] = quantity
            raw_price = ref.get("unit_price")
            unit_price = float(raw_price) if raw_price not in (None, "") else None
            row["unit_price"] = unit_price
            row["monthly_multiplier"] = self._monthly_multiplier_for_metric(str(row.get("metric") or ""))
            row["extended_price"] = (
                round(quantity * unit_price * float(row["monthly_multiplier"]), 4)
                if unit_price is not None else None
            )
            row["price_source"] = str(ref.get("source") or "unpriced_catalog")
            row["pricing_status"] = "priced" if unit_price is not None else "unpriced"
            output.append(row)

        sku_set = {str(row.get("sku") or "").strip().upper() for row in output}
        for cpu_sku, mem_sku in CPU_SKU_TO_MEM_SKU.items():
            if cpu_sku in sku_set and mem_sku not in sku_set:
                cpu_row = next((r for r in output if str(r.get("sku")).upper() == cpu_sku), None)
                qty = float(cpu_row.get("quantity", 1.0) if cpu_row else 1.0)
                mem_ref = pricing_table[mem_sku]
                output.append(
                    {
                        "sku": mem_sku,
                        "description": str(mem_ref.get("description") or mem_sku),
                        "category": "compute",
                        "metric": str(mem_ref.get("metric") or ""),
                        "quantity": max(1.0, qty * 16.0),
                        "unit_price": mem_ref.get("unit_price"),
                        "monthly_multiplier": self._monthly_multiplier_for_metric(str(mem_ref.get("metric") or "")),
                        "extended_price": round(
                            max(1.0, qty * 16.0)
                            * float(mem_ref.get("unit_price") or 0.0)
                            * self._monthly_multiplier_for_metric(str(mem_ref.get("metric") or "")),
                            4,
                        ) if mem_ref.get("unit_price") is not None else None,
                        "price_source": str(mem_ref.get("source") or "unpriced_catalog"),
                        "pricing_status": "priced" if mem_ref.get("unit_price") is not None else "unpriced",
                        "notes": "Auto-repair: added memory SKU for non-GPU split rule",
                    }
                )

        fixed["line_items"] = output
        return self._normalize_payload(fixed)

    def _normalize_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        norm = json.loads(json.dumps(payload))
        line_items = norm.get("line_items") or []
        total = 0.0
        priced_count = 0
        unpriced_count = 0
        for row in line_items:
            qty = float(row.get("quantity") or 0)
            raw_price = row.get("unit_price")
            price = float(raw_price) if raw_price not in (None, "") else None
            multiplier = float(row.get("monthly_multiplier") or self._monthly_multiplier_for_metric(str(row.get("metric") or "")))
            row["quantity"] = round(qty, 4)
            row["unit_price"] = round(price, 6) if price is not None else None
            row["monthly_multiplier"] = int(multiplier) if multiplier.is_integer() else round(multiplier, 4)
            row["extended_price"] = round(qty * price * multiplier, 4) if price is not None else None
            row["pricing_status"] = "priced" if price is not None else "unpriced"
            if price is None:
                unpriced_count += 1
            else:
                priced_count += 1
                total += float(row["extended_price"])
        norm["currency"] = str(norm.get("currency") or "USD")
        norm["monthly_total"] = round(total, 4)
        norm["totals"] = {"estimated_monthly_cost": round(total, 4)}
        norm["pricing_status"] = (
            "partially_priced" if priced_count and unpriced_count
            else "priced" if priced_count
            else "unpriced"
        )
        assumptions = norm.get("assumptions")
        if not isinstance(assumptions, list):
            norm["assumptions"] = []
        return norm

    def generate_xlsx(self, payload: dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # Columns: A=SKU, B=Description, C=Instance Count, D=Category,
        #          E=Metric, F=Quantity, G=Monthly Multiplier,
        #          H=Unit Price (USD), I=Monthly Cost (USD), J=Notes
        headers = [
            "SKU",
            "Description",
            "Instance Count",
            "Category",
            "Metric",
            "Quantity",
            "Monthly Multiplier",
            "Unit Price (USD)",
            "Monthly Cost (USD)",
            "Notes",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        line_items = payload.get("line_items") or []
        for item in line_items:
            raw_count = item.get("instance_count")
            count_val = int(raw_count) if raw_count not in (None, "", 0) else ""
            unit_price = item.get("unit_price")
            ws.append(
                [
                    item.get("sku", ""),
                    item.get("description", ""),
                    count_val,
                    item.get("category", ""),
                    item.get("metric", ""),
                    float(item.get("quantity") or 0),
                    float(item.get("monthly_multiplier") or 1),
                    float(unit_price) if unit_price not in (None, "") else "TBD",
                    None,  # formula column
                    item.get("notes", ""),
                ]
            )

        start_row = 2
        end_row = max(1, len(line_items) + 1)
        for row_idx in range(start_row, end_row + 1):
            if isinstance(ws.cell(row=row_idx, column=8).value, (int, float)):
                ws.cell(row=row_idx, column=9, value=f"=F{row_idx}*G{row_idx}*H{row_idx}")
            else:
                ws.cell(row=row_idx, column=9, value="TBD")

        total_row = end_row + 2
        ws.cell(row=total_row, column=8, value="TOTAL")
        ws.cell(row=total_row, column=9, value=f"=SUM(I{start_row}:I{end_row})")
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=9).font = Font(bold=True)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 32
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 18
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 42

        scope_items = [
            item for item in payload.get("scope_items", []) or [] if isinstance(item, dict)
        ]
        if scope_items:
            scope_ws = wb.create_sheet("Scope Coverage")
            scope_ws.append([
                "Canonical Service ID",
                "Description",
                "Quantity",
                "Pricing Status",
                "Assumed Sizing",
                "Assumptions",
                "Source",
            ])
            for cell in scope_ws[1]:
                cell.font = Font(bold=True)
            for item in scope_items:
                scope_ws.append([
                    item.get("canonical_service_id", ""),
                    item.get("description", ""),
                    item.get("quantity", ""),
                    item.get("pricing_status", ""),
                    json.dumps(item.get("assumed_sizing", {}), sort_keys=True),
                    "; ".join(str(value) for value in item.get("assumptions", []) or []),
                    item.get("source", ""),
                ])
            for column, width in zip("ABCDEFG", (34, 34, 12, 28, 42, 72, 20)):
                scope_ws.column_dimensions[column].width = width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


_SHARED_SERVICE: BomService | None = None


def get_shared_bom_service() -> BomService:
    global _SHARED_SERVICE
    if _SHARED_SERVICE is None:
        _SHARED_SERVICE = BomService()
    return _SHARED_SERVICE


def new_trace_id() -> str:
    return str(uuid.uuid4())
