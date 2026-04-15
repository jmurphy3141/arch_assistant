"""
agent/bom_parser.py
--------------------
Reads BOM Excel → produces a clean service list for the LLM layout compiler.
The LLM receives a structured summary and returns a hierarchical layout spec JSON.
This file does NOT decide layout — only service identification.

The LLM prompt uses an assumption-first approach: apply defaults from the
default assumption table, never ask about things that can be inferred.
"""
from __future__ import annotations
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ── SKU → (oci_type, default_layer) ─────────────────────────────────────────
SKU_MAP: dict[str, tuple] = {
    # Compute — VM shapes (all generations map to "compute")
    "B94176":  ("compute",           "compute"),   # E3/E4 OCPU
    "B94177":  (None,                None),        # E3/E4 memory — part of compute
    "B111129": ("compute",           "compute"),   # E6 OCPU
    "B111130": (None,                None),        # E6 memory — part of compute
    "B88317":  ("compute",           "compute"),   # A1 Flex OCPU
    "B88318":  (None,                None),        # A1 memory — part of compute
    # Block Storage
    "B91961":  (None,                None),        # block volume storage — implied by compute
    "B91962":  (None,                None),        # block volume perf — implied
    # Database
    "B99060":  ("database",          "data"),
    "B99062":  (None,                None),        # db storage — implied
    # Object Storage
    "B91628":  ("object storage",    "data"),
    # Load Balancer
    "B93030":  ("load balancer",     "ingress"),
    "B93031":  (None,                None),        # LB bandwidth — implied
    # FastConnect / DRG
    "B88325":  ("drg",               "ingress"),   # FastConnect 1 Gbps
    "B88326":  ("drg",               "ingress"),   # FastConnect 10 Gbps
    "B88327":  ("drg",               "ingress"),   # FastConnect 100 Gbps
    # Functions
    "B90618":  ("functions",         "compute"),
    "B90617":  (None,                None),        # fn execution — part of functions
    # API Gateway / Queue
    "B92072":  ("api gateway",       "ingress"),
    "B95697":  ("queue",             "async"),
}

DESC_MAP: dict[str, tuple] = {
    # ── Compute: Bare Metal (most specific first) ────────────────────────────
    "bm.optimized":           ("bare metal",       "compute"),  # HPC BM.Optimized3.36
    "bm.hpc":                 ("bare metal",       "compute"),
    "bm.standard":            ("bare metal",       "compute"),
    "bm.densio":              ("bare metal",       "compute"),
    "bm.gpu":                 ("bare metal",       "compute"),
    "bare metal":             ("bare metal",       "compute"),
    "rdma":                   ("bare metal",       "compute"),  # RDMA cluster

    # ── Compute: Container / OKE ─────────────────────────────────────────────
    "container instances":    ("container engine", "compute"),
    "oke enhanced":           ("container engine", "compute"),
    "oke":                    ("container engine", "compute"),
    "kubernetes":             ("container engine", "compute"),

    # ── Compute: VM shapes ───────────────────────────────────────────────────
    "vm.standard":            ("compute",          "compute"),
    "vm.optimized":           ("compute",          "compute"),
    "vm.gpu":                 ("compute",          "compute"),
    "gpu":                    ("compute",          "compute"),
    "flex":                   ("compute",          "compute"),  # E3/E4/E6/A1 Flex
    "standard - e":           ("compute",          "compute"),  # E3/E4/E6
    "standard - a":           ("compute",          "compute"),  # Ampere A1
    "ocpu per hour":          ("compute",          "compute"),  # any OCPU billing row

    # ── Compute: Analytics / Data platforms ──────────────────────────────────
    "data science":           ("compute",          "compute"),  # OCI Data Science notebooks
    "big data":               ("compute",          "compute"),  # BDS master/worker nodes
    "analytics cloud":        ("compute",          "compute"),  # OAC
    "data flow":              ("compute",          "compute"),  # Spark/Data Flow
    "goldengate":             ("compute",          "compute"),  # GoldenGate Microservices
    "data integration":       ("compute",          "compute"),  # OCI DIS
    "integration":            ("compute",          "compute"),  # Oracle Integration Cloud
    "visual builder":         ("compute",          "compute"),  # Oracle VBCS
    "process automation":     ("compute",          "compute"),  # OPA

    # ── Database ─────────────────────────────────────────────────────────────
    "autonomous database":    ("database",         "data"),
    "autonomous transaction":  ("database",        "data"),     # ATP
    "autonomous data warehouse": ("database",      "data"),     # ADW
    "heatwave":               ("database",         "data"),     # MySQL HeatWave
    "mysql":                  ("database",         "data"),
    "postgresql":             ("database",         "data"),
    "nosql":                  ("database",         "data"),
    "exadata":                ("database",         "data"),
    "base database":          ("database",         "data"),
    "database service":       ("database",         "data"),
    "opensearch":             ("database",         "data"),     # OCI Search/OpenSearch
    "cache":                  ("database",         "data"),     # OCI Cache (Redis)
    "redis":                  ("database",         "data"),

    # ── Storage ──────────────────────────────────────────────────────────────
    "file storage":           ("file storage",     "data"),
    "object storage":         ("object storage",   "data"),
    "block volume":           (None,               None),       # implied by compute — skip

    # ── Networking / Ingress ─────────────────────────────────────────────────
    "fastconnect":            ("drg",              "ingress"),
    "network load balancer":  ("load balancer",    "ingress"),
    "load balancer":          ("load balancer",    "ingress"),
    "api gateway":            ("api gateway",      "ingress"),
    "bastion":                ("bastion",          "ingress"),
    "web application firewall": ("waf",            "ingress"),
    "waf":                    ("waf",              "ingress"),

    # ── Async / Messaging / Integration ──────────────────────────────────────
    "streaming":              ("queue",            "async"),
    "queue":                  ("queue",            "async"),
    "kafka":                  ("queue",            "async"),
    "email delivery":         ("queue",            "async"),
    "notifications":          ("queue",            "async"),
    "events":                 ("queue",            "async"),
    "functions":              ("functions",        "compute"),

    # ── Security / IAM ───────────────────────────────────────────────────────
    "identity and access":    ("iam",              "data"),
    "vault":                  ("vault",            "data"),
    "secrets on oci vault":   ("vault",            "data"),
    "key management":         ("vault",            "data"),     # KMS
    "certificates":           ("vault",            "data"),
    "cloud guard":            ("monitoring",       "data"),
    "security advisor":       ("monitoring",       "data"),
    "vulnerability scanning": ("monitoring",       "data"),

    # ── Observability / Management ───────────────────────────────────────────
    "logging":                ("logging",          "data"),
    "monitoring":             ("monitoring",       "data"),
    "application performance monitoring": ("monitoring", "data"),
    "operations insights":    ("monitoring",       "data"),
    "database management":    ("monitoring",       "data"),
    "devops":                 ("monitoring",       "data"),     # OCI DevOps service

    # ── Explicit skips (billing rows that carry no architectural node) ────────
    "bandwidth":              (None,               None),
    "data transfer":          (None,               None),
    "dns":                    (None,               None),
    "virtual cloud network":  (None,               None),
    "vcn":                    (None,               None),
    "support":                (None,               None),
    "overage":                (None,               None),
    "commitment":             (None,               None),
}


# ── Description normalisation ────────────────────────────────────────────────

_OCI_PREFIXES = (
    "oracle cloud infrastructure ",
    "oracle cloud ",
    "oci ",
    "oracle ",
)

def _normalize_desc(desc: str) -> str:
    """Strip Oracle/OCI branding so 'Oracle Cloud Infrastructure MySQL ...' matches 'mysql'."""
    for prefix in _OCI_PREFIXES:
        if desc.startswith(prefix):
            return desc[len(prefix):].strip()
    return desc


# ── Token-based last-resort inference ────────────────────────────────────────

# Billing/infra tokens that produce no architectural node
_SKIP_TOKENS = frozenset({
    "bandwidth", "data transfer", "transfer", "license", "support",
    "overage", "commitment", "prepay", "vcn", "subnet", "vnic",
    "dhcp", "route table", "security list", "nsg", "peering",
    "reserved ip", "private ip", "nat ip",
})

def _infer_from_tokens(desc: str) -> tuple | None:
    """
    Token-level heuristics for OCI billing descriptions that didn't match DESC_MAP.

    OCI billing rows reliably encode their category via billing unit:
      - "OCPU" / "OCPU per hour"       → compute (or database when combined with ADB keywords)
      - "ECPU" / "ECPU per hour"       → database (only ADB/ADB-S uses ECPU billing)
      - "storage", "gb per month"      → None (implied, skip)
      - known skip tokens              → None (billing overhead, skip)

    Returns (oci_type, layer), (None, None) to skip, or None when confidence is low.
    """
    # Hard skip: known billing/infra overhead tokens
    if any(token in desc for token in _SKIP_TOKENS):
        return (None, None)

    # ECPU rows are exclusively Autonomous Database billing
    if "ecpu" in desc:
        return ("database", "data")

    # OCPU rows: ADB OCPU → database; everything else → compute
    if "ocpu" in desc:
        if any(db in desc for db in ("autonomous", "adb", "data warehouse", "transaction processing")):
            return ("database", "data")
        return ("compute", "compute")

    # Node rows in cluster contexts
    if "node" in desc and any(t in desc for t in ("worker", "master", "infra", "compute")):
        return ("compute", "compute")

    # Storage rows: skip
    if any(t in desc for t in ("storage", "gb per month", "tb per month")):
        return (None, None)

    return None  # genuinely unknown — caller logs warning

# Best-practice additions (always injected)
BEST_PRACTICE = [
    {"id": "internet_gateway", "type": "internet gateway", "label": "Internet Gateway",     "layer": "external"},
    {"id": "nat_gateway",      "type": "nat gateway",      "label": "NAT Gateway",           "layer": "ingress"},
    {"id": "service_gateway",  "type": "service gateway",  "label": "Service Gateway",       "layer": "ingress"},
    {"id": "waf",              "type": "waf",              "label": "WAF",                   "layer": "ingress"},
    {"id": "network_firewall", "type": "waf",              "label": "Network Firewall",      "layer": "ingress"},
    {"id": "logging",          "type": "logging",          "label": "Logging Analytics\n(SIEM)", "layer": "data"},
    {"id": "monitoring",       "type": "monitoring",       "label": "Monitoring + APM",      "layer": "data"},
    {"id": "db_mgmt",          "type": "monitoring",       "label": "DB Management",         "layer": "data"},
    {"id": "directory",        "type": "iam",              "label": "Directory Services",    "layer": "data"},
    {"id": "certificates",     "type": "vault",            "label": "Certificates",          "layer": "data"},
]


@dataclass
class ServiceItem:
    id:       str
    oci_type: str
    label:    str
    layer:    str           # external | ingress | compute | async | data
    quantity: Optional[float] = None
    notes:    str = ""


# ── Multi-environment sheet detection ────────────────────────────────────────

# (priority, keyword) — higher number = more "production-like"
_ENV_PRIORITY: list[tuple[int, str]] = [
    (10, "prod"), (10, "production"),
    (8,  "prototype"), (8, "proto"),
    (5,  "dev"), (5, "development"),
    (3,  "pristine"), (3, "base"),
    (2,  "test"), (2, "uat"), (2, "staging"), (2, "preprod"),
]

def _env_sheet_priority(name: str) -> int:
    """Return priority for an environment-named sheet (0 = not an env sheet)."""
    n = name.lower()
    for priority, kw in _ENV_PRIORITY:
        if kw in n:
            return priority
    return 0


def _find_bom_sheets(wb) -> list[str]:
    """Return env sheet names sorted by priority (highest first).
    Falls back to 'BOM' sheet or first sheet if no env-named sheets found."""
    env_sheets = [(name, _env_sheet_priority(name)) for name in wb.sheetnames
                  if _env_sheet_priority(name) > 0]
    if env_sheets:
        env_sheets.sort(key=lambda x: -x[1])
        return [s[0] for s in env_sheets]
    if "BOM" in wb.sheetnames:
        return ["BOM"]
    return [wb.worksheets[0].title]


def _extract_sheet_quantities(sheet) -> dict:
    """Scan a BOM sheet and return aggregate quantities by service category."""
    hdrs = None
    compute_ocpu = 0.0
    db_qty = 0.0
    storage_gb = 0.0
    for row in sheet.iter_rows(values_only=True):
        if hdrs is None:
            candidate = [str(c).lower().strip() if c else "" for c in row]
            if "sku" in candidate and "description" in candidate:
                hdrs = candidate
            continue
        if not any(v is not None for v in row):
            continue
        d = dict(zip(hdrs, row))
        desc = str(d.get("description", "") or "").lower()
        qty = float(d.get("quantity") or 0)
        if "ocpu" in desc and not any(t in desc for t in ("autonomous", "adb")):
            compute_ocpu += qty
        elif any(t in desc for t in ("autonomous", "adb", "ecpu")):
            db_qty += qty
        elif any(t in desc for t in ("storage", "gb per month", "tb per month", "block volume")):
            storage_gb += qty
    return {"compute_ocpu": int(compute_ocpu), "db_qty": int(db_qty), "storage_gb": int(storage_gb)}


def _build_env_summary(wb, env_sheets: list[str]) -> str:
    """Build a text summary of per-environment resource quantities for the LLM prompt."""
    lines = [
        "MULTI-ENVIRONMENT BOM: This workbook contains separate environment tabs.",
        "The architecture diagram is drawn from the primary environment below.",
        "Other environments use the SAME topology with different resource sizes.",
        "",
        "Environments detected:",
    ]
    for sheet_name in env_sheets:
        q = _extract_sheet_quantities(wb[sheet_name])
        lines.append(
            f"  • {sheet_name}: {q['compute_ocpu']} compute OCPUs"
            + (f", {q['db_qty']} DB OCPU/ECPU" if q["db_qty"] else "")
            + (f", {q['storage_gb']:.0f} GB storage" if q["storage_gb"] else "")
        )
    lines.append(f"\nPrimary environment (used for diagram): {env_sheets[0]}")
    return "\n".join(lines)


# Types that represent shared infrastructure — only kept from the primary section.
# Secondary environments (DR, Dev, etc.) contribute workload types only.
_SHARED_INFRA_TYPES = frozenset([
    "internet gateway", "nat gateway", "service gateway", "drg",
    "waf", "load balancer", "logging", "monitoring", "iam", "vault",
    "internet", "on premises", "bastion",
])


def _split_bom_into_sections(sheet) -> list[tuple[str, list, list]]:
    """Split a single BOM sheet into named environment sections.

    Each section starts with an optional label row (e.g. "Prod", "DR") followed
    by a header row that contains both "sku" and "description" columns, then the
    data rows for that environment.

    Returns: list of (section_name, col_headers, row_dicts)
    """
    result: list[tuple[str, list, list]] = []
    next_name: str = "Primary"   # name queued for the next section
    cur_name:  str | None = None
    cur_hdrs:  list | None = None
    cur_rows:  list = []

    for row in sheet.iter_rows(values_only=True):
        # Skip fully empty rows
        if not any(v is not None and str(v).strip() for v in row):
            continue

        col_names = [str(c).lower().strip() if c else "" for c in row]

        # ── Is this a column-header row? ─────────────────────────────────────
        if "sku" in col_names and "description" in col_names:
            # Flush the current section before starting a new one
            if cur_hdrs is not None and cur_rows:
                result.append((cur_name, cur_hdrs, cur_rows))
            cur_name  = next_name
            next_name = f"Section {len(result) + 2}"  # fallback if no label precedes next
            cur_hdrs  = col_names
            cur_rows  = []
            continue

        # ── Is this a section-label row? ─────────────────────────────────────
        # A label row has ≤2 non-empty cells, and the first cell is short text
        # that doesn't look like a SKU or a plain number.
        # Threshold is 2 (not 3) so that standard 3-column data rows
        # (SKU, Description, Quantity) are never mis-classified as labels,
        # even when the SKU doesn't match the canonical B+digits OCI pattern.
        text_cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if text_cells:
            first = text_cells[0]
            is_sku = (
                # Standard OCI SKU: B followed by digits only (e.g. B94176)
                (len(first) >= 5 and first[0].upper() == "B" and first[1:].isdigit())
                # Any all-caps identifier without spaces (e.g. ZZZUNKNOWN, BFUTURE1, BBANDWIDTH)
                or (len(first) >= 5 and " " not in first and first.upper() == first)
            )
            is_number = first.replace(".", "").replace(",", "").isdigit()
            if len(text_cells) <= 2 and not is_sku and not is_number and len(first) < 60:
                next_name = first
                continue

        # ── Regular data row ──────────────────────────────────────────────────
        if cur_hdrs is not None:
            cur_rows.append(dict(zip(cur_hdrs, row)))

    # Flush last section
    if cur_hdrs is not None and cur_rows:
        result.append((cur_name, cur_hdrs, cur_rows))

    return result


def _lookup_row(sku: str, desc: str) -> tuple[str | None, str | None]:
    """Tier 1→3 service-type lookup for a single BOM row.
    Returns (oci_type, layer); (None, None) means silently skip."""
    # Tier 1: exact SKU
    if sku in SKU_MAP:
        return SKU_MAP[sku]

    # Tier 2a: DESC_MAP on raw description
    for key, (t, l) in DESC_MAP.items():
        if key in desc:
            return t, l

    # Tier 2b: DESC_MAP on normalised description (strip OCI/Oracle prefix)
    norm = _normalize_desc(desc)
    if norm != desc:
        for key, (t, l) in DESC_MAP.items():
            if key in norm:
                return t, l

    # Tier 3: token-level billing-unit inference
    inferred = _infer_from_tokens(desc)
    if inferred is not None:
        return inferred

    return None, None   # genuinely unknown


def parse_bom(xlsx_path: str | Path, context: str = "",
              sheet_name: str | None = None) -> list[ServiceItem]:
    """Parse BOM Excel → list of ServiceItems ready for LLM layout prompt.

    Handles both single-environment and multi-environment BOMs.  A multi-
    environment BOM has repeated blocks in one sheet:
        Prod
        SKU | Description | Quantity ...
        <data rows>
        DR
        SKU | Description | Quantity ...
        <data rows>

    Each environment gets its own ServiceItems (de-duplicated within the
    section).  Shared infrastructure (gateways, WAF, etc.) is kept only from
    the primary section.  Secondary environment items are tagged with the
    section name in their ID and label.

    sheet_name: if given, force that specific sheet; otherwise auto-detect.
    """
    import openpyxl
    wb  = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        bom = wb[sheet_name]
    else:
        primary = _find_bom_sheets(wb)[0]
        bom = wb[primary]

    inp = wb["Input"] if "Input" in wb.sheetnames else None

    # Optional AWS-style Input sheet (usually absent in OCI BOMs)
    input_data: dict[str, dict] = {}
    if inp:
        hdrs = None
        for row in inp.iter_rows(values_only=True):
            if hdrs is None:
                hdrs = [str(c).lower() if c else "" for c in row]
                continue
            if row[0]:
                input_data[str(row[0]).lower()] = {
                    hdrs[i]: row[i] for i in range(1, len(row)) if i < len(hdrs)
                }

    app_ocpu = int((input_data.get("ec2", {}).get("vcpu count", 0) or 0) / 2)
    db_ocpu  = int((input_data.get("postgres rds", {}).get("vcpu count", 0) or 0) / 2)
    obj_gb   = input_data.get("postgres rds", {}).get("storage (gb)", 0) or 0

    # ── Split sheet into environment sections ─────────────────────────────────
    sections = _split_bom_into_sections(bom)
    if not sections:
        sections = [("Primary", [], [])]

    items:      list[ServiceItem] = []
    seen_global: set[str] = set()   # all oci_types seen so far (for infra dedup)

    for sec_idx, (sec_name, sec_hdrs, sec_rows) in enumerate(sections):
        is_primary = (sec_idx == 0)
        # Secondary sections prefix IDs and suffix labels with env name
        id_prefix  = "" if is_primary else f"{sec_name.lower()[:8].replace(' ', '_')}_"
        lbl_suffix = "" if is_primary else f"\n({sec_name})"

        seen_local:    set[str]       = set()
        counters_local: dict[str, int] = {}

        for d in sec_rows:
            sku  = str(d.get("sku", "") or "").strip()
            desc = str(d.get("description", "") or "").lower().strip()
            qty  = d.get("quantity")
            note = str(d.get(sec_hdrs[-1], "") or "") if sec_hdrs else ""

            oci_type, layer = _lookup_row(sku, desc)

            if oci_type is None and layer is None:
                # Genuinely unknown
                logger.warning(
                    "parse_bom [%s]: unrecognized SKU=%r desc=%r — row skipped",
                    sec_name, sku, desc,
                )
                continue
            if not oci_type:
                # (None, None) = explicit skip (billing overhead row)
                continue

            # Shared infrastructure only appears once (from the primary section)
            if not is_primary and oci_type in _SHARED_INFRA_TYPES:
                continue

            # Deduplicate within this section
            if oci_type in seen_local:
                continue
            seen_local.add(oci_type)
            seen_global.add(oci_type)

            counters_local[oci_type] = counters_local.get(oci_type, 0) + 1
            nid   = f"{id_prefix}{oci_type.replace(' ', '_')}_{counters_local[oci_type]}"
            label = _make_label(oci_type, qty, app_ocpu, db_ocpu, obj_gb, note) + lbl_suffix

            items.append(ServiceItem(id=nid, oci_type=oci_type, label=label,
                                     layer=layer, quantity=qty, notes=sec_name))

    # ── On-Premises (always) ──────────────────────────────────────────────────
    items.insert(0, ServiceItem(id="on_prem", oci_type="on premises",
                                label="On-Premises\n(3 Offices)", layer="external"))
    seen_global.add("on premises")

    # ── Best-practice services (once) ────────────────────────────────────────
    for bp in BEST_PRACTICE:
        if bp["type"] not in seen_global:
            items.append(ServiceItem(id=bp["id"], oci_type=bp["type"],
                                     label=bp["label"], layer=bp["layer"],
                                     notes="best practice"))
            seen_global.add(bp["type"])

    # ── Baseline injection: Internet ─────────────────────────────────────────
    if (
        "internet gateway" in seen_global
        and "NO_INTERNET_ENDPOINT=true" not in context
        and "internet" not in seen_global
    ):
        items.append(ServiceItem(id="internet", oci_type="internet",
                                 label="Public Internet", layer="external",
                                 notes="injected_baseline"))
        seen_global.add("internet")

    # ── Baseline injection: Bastion ──────────────────────────────────────────
    has_workload = any(i.oci_type in {"compute", "database"} for i in items)
    if (
        has_workload
        and "NO_BASTION=true" not in context
        and "bastion" not in seen_global
    ):
        items.append(ServiceItem(id="bastion_1", oci_type="bastion",
                                 label="Bastion", layer="ingress",
                                 notes="injected_baseline"))
        seen_global.add("bastion")

    return items


def _make_label(oci_type: str, qty, app_ocpu: int, db_ocpu: int, obj_gb: float, note: str) -> str:
    q = int(qty) if qty else "?"
    labels = {
        "compute":        f"Compute\n×{q} OCPU",
        "bare metal":     f"HPC BM.Optimized3.36\n×{q} nodes",
        "database":       f"PostgreSQL DB\n×{db_ocpu:,} OCPU" if db_ocpu else f"Database\n×{q} OCPU",
        "object storage": f"Object Storage\n{int(obj_gb*2/1024)} TB",
        "load balancer":  f"Load Balancer\n×{q} (per region)",
        "drg":            f"DRG / FastConnect\n×{q} ports",
        "functions":      "OCI Functions\n~10k calls/day",
        "api gateway":    "API Gateway",
        "queue":          f"Queue\n{q}M req/month",
        "container engine": "OKE Enhanced Cluster",
        "file storage":   "File Storage\nNFS PVC",
        "bastion":        "Bastion",
        "iam":            "IAM",
        "load balancer_2": "NLB ×3",
        "vault":          "Vault (Secrets)",
    }
    return labels.get(oci_type, oci_type.title())


def build_llm_prompt(items: list[ServiceItem], context: str = "") -> str:
    """Build the layout compiler prompt for the OCI GenAI agent.

    Uses an assumption-first approach: the LLM must apply the default assumption
    table and NEVER ask about things it can infer. Clarification is only for
    truly blocking topology decisions.

    context: optional text from a requirements/notes file uploaded alongside the BOM.

    Returns a prompt that instructs the LLM to output the new hierarchical spec:
    regions → availability_domains → fault_domains → subnets → nodes.
    """
    service_list = "\n".join(
        f'  {{"id": "{i.id}", "type": "{i.oci_type}", "label": "{i.label.replace(chr(10), " ")}", "suggested_layer": "{i.layer}"}}'
        for i in items
    )

    context_block = (
        f"\nADDITIONAL CONTEXT:\n{context.strip()}\n"
        if context and context.strip() else ""
    )

    return f"""You are an OCI architecture layout compiler. Your job is to produce a deterministic
hierarchical layout specification JSON for an OCI draw.io diagram.
{context_block}
═══════════════════════════════════════════════════════
ASSUMPTION-FIRST RULE (CRITICAL)
═══════════════════════════════════════════════════════
Apply the default assumption table below for any missing information.
NEVER ask about things you can infer. Only ask clarification questions for
information that would materially change the topology AND cannot be safely
assumed from context.

DEFAULT ASSUMPTION TABLE:
┌─────────────────────────────────────────────┬──────────────────────────────────────────────┐
│ Signal in BOM / context                     │ Assumed topology                             │
├─────────────────────────────────────────────┼──────────────────────────────────────────────┤
│ No HA signal at all                         │ single_ad, single FD                         │
│ "HA" or redundancy mentioned                │ single_ad, two Fault Domains                 │
│ Two ADs mentioned, or "regional HA"         │ multi_ad, two ADs side by side               │
│ "DR", "multi-region", or two regions        │ multi_region, active-passive                 │
│ "active-active" explicit                    │ single_ad, two Fault Domains (active-active) │
│ Database in BOM                             │ Add Data Guard (sync multi_ad, async multi_region) │
│ Any compute in BOM                          │ Add NAT Gateway                              │
│ Any OCI managed service                     │ Add Service Gateway                          │
│ External users / HTTPS in BOM or context    │ Add IGW + WAF + Public Load Balancer         │
│ On-prem / FastConnect in BOM                │ Add DRG + Private Load Balancer              │
│ No load balancer listed                     │ Add one (public or private per other signals)│
└─────────────────────────────────────────────┴──────────────────────────────────────────────┘

NEVER ASK ABOUT:
- Whether to include gateways (always add the appropriate ones based on signals above)
- Whether to include WAF (always add when internet-facing)
- Subnet count or naming (derive from tier model below)
- Icon styles or colours (always use OCI standards)
- Page size or layout direction (always A3 landscape 1654×1169, always TB)

═══════════════════════════════════════════════════════
INPUT SERVICES (from BOM):
═══════════════════════════════════════════════════════
[
{service_list}
]

═══════════════════════════════════════════════════════
LAYOUT RULES
═══════════════════════════════════════════════════════
Layout direction: TOP → BOTTOM (TB)
Canvas: 1654 × 1169 px (A3 landscape)

SUBNET TIER MODEL (top to bottom inside each AD):
  ingress — Public Subnet: WAF, Public Load Balancer, Bastion
  ingress — Private Subnet: Private Load Balancer, DRG connectivity
  web     — Private Subnet: Web Tier compute
  app     — Private Subnet: App Tier compute, Functions, API Gateway, Queues
  db      — Private Subnet: Databases, Vault

PLACEMENT RULES:
1. Regional subnets (LB, Bastion, WAF) are placed ABOVE the AD boxes, inside the region.
2. Gateways straddle the region box edges:
   - internet_gateway → top edge (position: "top")
   - drg              → left edge (position: "left")
   - nat_gateway      → right edge (position: "right")
   - service_gateway  → right edge, below NAT (position: "right")
3. OCI managed services (Object Storage, IAM, Logging, Monitoring) → oci_services list.
4. External elements (on-premises, internet users, CPE) → external list.
   - id="internet" (type "internet") MUST appear in external[].
5. For single_ad: include fault_domains[] inside the AD; subnets at AD level = shared tiers (DB).
6. For multi_ad / multi_region: no fault_domains — subnets directly inside each AD.
7. id="bastion_1" (type "bastion") MUST be placed inside a regional_subnets[] ingress subnet.

═══════════════════════════════════════════════════════
CLARIFICATION (ONLY IF TRULY BLOCKING)
═══════════════════════════════════════════════════════
If — and ONLY if — there is a specific topology decision that cannot be determined
from the BOM, context, or default assumption table, return ONLY:
{{
  "status": "need_clarification",
  "questions": ["<single concise question>"]
}}
Keep it to at most one or two truly blocking questions. Do NOT ask about gateways,
WAF, subnet naming, icon styles, page size, or anything in the assumption table.

═══════════════════════════════════════════════════════
OUTPUT JSON SCHEMA (use this exact structure)
═══════════════════════════════════════════════════════
{{
  "deployment_type": "single_ad",
  "page": {{"width": 1654, "height": 1169}},
  "regions": [
    {{
      "id": "region_primary",
      "label": "Oracle Cloud Infrastructure (Region)",
      "regional_subnets": [
        {{
          "id": "pub_sub_lb",
          "label": "Public Subnet",
          "tier": "ingress",
          "nodes": [
            {{"id": "waf_1",    "type": "waf",           "label": "WAF"}},
            {{"id": "pub_lb_1", "type": "load balancer", "label": "Load Balancer"}}
          ]
        }},
        {{
          "id": "pub_sub_bastion",
          "label": "Public Subnet",
          "tier": "ingress",
          "nodes": [{{"id": "bastion_1", "type": "bastion", "label": "Bastion Host"}}]
        }}
      ],
      "availability_domains": [
        {{
          "id": "ad1",
          "label": "Availability Domain 1",
          "fault_domains": [
            {{
              "id": "fd1",
              "label": "Fault Domain 1",
              "subnets": [
                {{
                  "id": "web_sub_fd1",
                  "label": "Private Subnet",
                  "tier": "web",
                  "nodes": [{{"id": "web_1", "type": "compute", "label": "Web Tier"}}]
                }},
                {{
                  "id": "app_sub_fd1",
                  "label": "Private Subnet",
                  "tier": "app",
                  "nodes": [{{"id": "app_1", "type": "compute", "label": "App Tier"}}]
                }}
              ]
            }}
          ],
          "subnets": [
            {{
              "id": "db_sub",
              "label": "Private Subnet",
              "tier": "db",
              "nodes": [{{"id": "db_1", "type": "database", "label": "PostgreSQL DB"}}]
            }}
          ]
        }}
      ],
      "gateways": [
        {{"id": "igw_1",  "type": "internet gateway", "label": "Internet Gateway", "position": "top"}},
        {{"id": "drg_1",  "type": "drg",              "label": "DRG",              "position": "left"}},
        {{"id": "nat_1",  "type": "nat gateway",      "label": "NAT Gateway",      "position": "right"}},
        {{"id": "sgw_1",  "type": "service gateway",  "label": "Service Gateway",  "position": "right"}}
      ],
      "oci_services": [
        {{"id": "obj_storage_1", "type": "object storage", "label": "Object Storage"}},
        {{"id": "logging_1",     "type": "logging",         "label": "Logging Analytics"}}
      ]
    }}
  ],
  "external": [
    {{"id": "on_prem",  "type": "on premises",  "label": "On-Premises"}},
    {{"id": "internet", "type": "internet",      "label": "Public Internet"}},
    {{"id": "admins",   "type": "users",         "label": "Admins"}}
  ],
  "edges": [
    {{"id": "e1", "source": "on_prem",   "target": "drg_1",    "label": "FastConnect"}},
    {{"id": "e2", "source": "internet",  "target": "igw_1",    "label": "HTTPS/443"}},
    {{"id": "e3", "source": "igw_1",     "target": "waf_1",    "label": "HTTPS/443"}},
    {{"id": "e4", "source": "waf_1",     "target": "pub_lb_1", "label": "HTTPS/443"}},
    {{"id": "e5", "source": "nat_1",     "target": "internet", "label": "Outbound"}}
  ]
}}

IMPORTANT RULES FOR OUTPUT:
1. Use ONLY services from the INPUT SERVICES list above. Do not invent new services.
2. Every service from the INPUT list must appear exactly once in the output.
3. Assign IDs exactly as given in the INPUT (e.g. "compute_1", "database_1").
4. deployment_type must be one of: "single_ad", "multi_ad", "multi_region".
5. For multi_ad: two availability_domains side by side, no fault_domains.
6. For multi_region: two entries in regions[], each a full single-AD layout.
7. Apply the default assumption table — do NOT ask if you can infer.
8. Output ONLY valid JSON. No markdown, no prose, no code fences."""


# ── Topology pattern detection ────────────────────────────────────────────────

_TOPOLOGY_HPC_OKE       = "HPC_OKE"
_TOPOLOGY_DATA_PLATFORM = "DATA_PLATFORM"
_TOPOLOGY_3TIER         = "STANDARD_3TIER"

# Pre-declared groups per topology (injected as directives, not suggestions)
_TOPOLOGY_GROUPS: dict[str, list[dict]] = {
    _TOPOLOGY_HPC_OKE: [
        {"id": "bas_sub_box",    "label": "Bastion Subnet (Public)", "order": 0},
        {"id": "cp_sub_box",     "label": "Control Plane Subnet",    "order": 1},
        {"id": "worker_sub_box", "label": "Worker Subnet (Private)", "order": 2},
        {"id": "storage_sub_box","label": "Storage Subnet",          "order": 3},
    ],
    _TOPOLOGY_DATA_PLATFORM: [
        {"id": "ingest_sub_box", "label": "Ingest Subnet",    "order": 0},
        {"id": "proc_sub_box",   "label": "Processing Subnet","order": 1},
        {"id": "store_sub_box",  "label": "Storage Subnet",   "order": 2},
    ],
    _TOPOLOGY_3TIER: [
        {"id": "pub_sub_box",    "label": "Public Subnet",    "order": 0},
        {"id": "app_sub_box",    "label": "App Subnet",       "order": 1},
        {"id": "db_sub_box",     "label": "DB Subnet",        "order": 2},
    ],
}

# Explicit group assignment per oci_type per topology
_TOPOLOGY_GROUP_MAP: dict[str, dict[str, str]] = {
    _TOPOLOGY_HPC_OKE: {
        "bastion":          "bas_sub_box",
        "waf":              "bas_sub_box",
        "load balancer":    "bas_sub_box",
        "container engine": "cp_sub_box",
        "bare metal":       "worker_sub_box",
        "compute":          "worker_sub_box",
        "file storage":     "storage_sub_box",
        "database":         "storage_sub_box",
        "vault":            "storage_sub_box",
    },
    _TOPOLOGY_DATA_PLATFORM: {
        "waf":              "ingest_sub_box",
        "load balancer":    "ingest_sub_box",
        "bastion":          "ingest_sub_box",
        "api gateway":      "ingest_sub_box",
        "queue":            "ingest_sub_box",
        "functions":        "proc_sub_box",
        "compute":          "proc_sub_box",
        "container engine": "proc_sub_box",
        "database":         "store_sub_box",
        "file storage":     "store_sub_box",
        "vault":            "store_sub_box",
    },
    _TOPOLOGY_3TIER: {
        "waf":              "pub_sub_box",
        "load balancer":    "pub_sub_box",
        "bastion":          "pub_sub_box",
        "compute":          "app_sub_box",
        "container engine": "app_sub_box",
        "functions":        "app_sub_box",
        "api gateway":      "app_sub_box",
        "queue":            "app_sub_box",
        "database":         "db_sub_box",
        "vault":            "db_sub_box",
        "file storage":     "db_sub_box",
    },
}

# oci_types that are never placed inside a subnet group (gateways + managed services)
_NO_GROUP_TYPES = frozenset([
    "internet gateway", "nat gateway", "service gateway", "drg",
    "object storage", "logging", "monitoring", "iam", "certificates",
    "on premises", "internet", "users", "admins", "workstation",
])


def _detect_topology_pattern(items: list[ServiceItem]) -> str:
    """Classify the architecture pattern from service types (deterministic, no LLM)."""
    types = {i.oci_type for i in items}
    has_bare_metal = "bare metal" in types
    has_oke        = "container engine" in types
    has_fss        = "file storage" in types

    # Bare metal alone or OKE + FSS both signal HPC workloads
    if has_bare_metal or (has_oke and has_fss):
        return _TOPOLOGY_HPC_OKE

    # Streaming-dominated, no bare metal, no significant relational DB
    has_streaming = "queue" in types or "streaming" in types
    has_db        = "database" in types
    if has_streaming and not has_bare_metal and not has_db:
        return _TOPOLOGY_DATA_PLATFORM

    return _TOPOLOGY_3TIER


def _build_edge_examples(items: list[ServiceItem], topology: str) -> str:
    """Generate topology-appropriate edge examples using actual service IDs from the BOM."""
    import json as _json
    ids = {i.oci_type: i.id for i in items}  # last id wins (fine for examples)

    edges: list[dict] = []
    ctr = [0]

    def _e(src_type: str, tgt_type: str, label: str) -> None:
        src = ids.get(src_type)
        tgt = ids.get(tgt_type)
        if src and tgt and src != tgt:
            ctr[0] += 1
            edges.append({"id": f"e{ctr[0]}", "source": src, "target": tgt, "label": label})

    if topology == _TOPOLOGY_HPC_OKE:
        # Cross-boundary entry (non-obvious: requires explicit FastConnect link)
        _e("on premises",      "drg",           "FastConnect")
        # Special RDMA fabric between compute nodes (non-obvious high-speed network)
        _e("container engine", "bare metal",    "RDMA")
        # Storage mount (non-obvious: NFS over dedicated storage network)
        _e("bare metal",       "file storage",  "NFS")

    elif topology == _TOPOLOGY_DATA_PLATFORM:
        # Application data-plane paths (non-obvious service-to-service calls)
        _e("api gateway",      "functions",      "Invoke")
        _e("functions",        "queue",          "Enqueue")
        _e("functions",        "database",       "SQL")
        _e("functions",        "object storage", "PUT/GET")

    else:  # STANDARD_3TIER
        # Cross-boundary entry
        _e("on premises",   "drg",          "FastConnect")
        # Application data path (non-obvious tier-to-tier calls)
        _e("load balancer", "compute",      "HTTP")
        _e("compute",       "database",     "SQL/5432")

    return _json.dumps(edges, indent=4)


def build_layout_intent_prompt(
    items: list[ServiceItem],
    questionnaire_text: str = "",
    notes_text: str = "",
    context: str = "",
) -> str:
    """
    Build a compact LayoutIntent prompt.

    Python detects the likely topology from BOM service types and injects it
    as a suggestion.  The LLM owns topology detection and can override based on
    free-text context.  The LLM also declares data-flow edges between services —
    that is the primary value it adds over deterministic Python logic.

    questionnaire_text: answers to a pre-flight questionnaire, if any.
    notes_text:         meeting notes or other free-text input, if any.
    context:            generic context string (e.g. from uploaded context file).
    """
    import json as _json

    topology   = _detect_topology_pattern(items)
    groups     = _TOPOLOGY_GROUPS[topology]
    group_map  = _TOPOLOGY_GROUP_MAP[topology]

    topology_label = {
        _TOPOLOGY_HPC_OKE:       "HPC on OKE (bare metal RDMA + container engine + file storage)",
        _TOPOLOGY_DATA_PLATFORM: "Data Platform (streaming + functions + data lake)",
        _TOPOLOGY_3TIER:         "Standard 3-Tier (web app + compute + database)",
    }[topology]

    # Service list with group hints (suggestions, not mandates)
    def _group_hint(item: ServiceItem) -> str:
        if item.oci_type in _NO_GROUP_TYPES:
            return "null"
        hint = group_map.get(item.oci_type)
        return f'"{hint}"' if hint else "null"

    service_rows = "\n".join(
        f'  {{"id": "{i.id}", "oci_type": "{i.oci_type}", "suggested_layer": "{i.layer}", "group_hint": {_group_hint(i)}}}'
        for i in items
    )

    suggested_groups_json = _json.dumps(groups, indent=4)
    example_groups_json   = _json.dumps(
        [{"id": g["id"], "label": g["label"], "order": g["order"]} for g in groups],
        indent=4,
    )
    edge_examples = _build_edge_examples(items, topology)

    # ── Multi-environment detection ──────────────────────────────────────────
    # Items tagged with environment names (notes field) when BOM has multiple sections.
    _reserved_notes = {"best practice", "injected_baseline", ""}
    env_names = list(dict.fromkeys(
        i.notes for i in items
        if i.notes not in _reserved_notes
    ))
    multi_env_block = ""
    if len(env_names) > 1:
        env_items_by_name: dict[str, list[ServiceItem]] = {}
        for i in items:
            if i.notes not in _reserved_notes:
                env_items_by_name.setdefault(i.notes, []).append(i)
        env_lines = []
        for env, env_items in env_items_by_name.items():
            types = ", ".join(sorted({it.oci_type for it in env_items
                                      if it.oci_type not in _SHARED_INFRA_TYPES}))
            env_lines.append(f"  • {env}: {types}")
        multi_env_block = (
            "\nMULTI-ENVIRONMENT BOM:\n"
            f"This BOM contains {len(env_names)} environments: {', '.join(env_names)}.\n"
            "The INPUT SERVICES list below includes separate items for each environment\n"
            "(item IDs are prefixed with the environment name for non-primary environments).\n"
            "Shared infrastructure (gateways, WAF, monitoring) appears only once.\n"
            "\nWorkloads per environment:\n" + "\n".join(env_lines) + "\n"
            "\nRECOMMENDED LAYOUT: Create one subnet group per environment for workload\n"
            "services (e.g. 'Prod App Subnet', 'DR App Subnet'). Shared infra sits\n"
            "outside environment groups as usual.\n"
        )

    extra_blocks = ""
    if multi_env_block:
        extra_blocks += multi_env_block
    if questionnaire_text and questionnaire_text.strip():
        extra_blocks += f"\nQUESTIONNAIRE ANSWERS:\n{questionnaire_text.strip()}\n"
    if notes_text and notes_text.strip():
        extra_blocks += f"\nMEETING / NOTES:\n{notes_text.strip()}\n"
    if context and context.strip():
        extra_blocks += f"\nADDITIONAL CONTEXT:\n{context.strip()}\n"

    return f"""You are an OCI solutions architect and layout compiler.
Given a list of OCI services from a Bill of Materials, you must:
  1. Decide the subnet topology (groups)
  2. Assign each service to the correct layer and group
  3. Declare the data-flow edges between services
  4. Set deployment hints (region count, HA, DR, connectivity)
Output ONLY valid JSON — either a LayoutIntent or a NeedClarification object.
{extra_blocks}
═══════════════════════════════════════════════════════
SUGGESTED TOPOLOGY (detected from BOM service types):
  {topology_label}
═══════════════════════════════════════════════════════
This is a suggestion based on what services are present.
If ADDITIONAL CONTEXT above indicates a different architecture, use that instead.

Suggested groups (use these unless context requires a different topology):
{suggested_groups_json}

Suggested group assignment per service type:
{chr(10).join(f"  {t!r:30s} → {g}" for t, g in group_map.items())}
  — gateways (internet gateway, nat gateway, service gateway, drg) → group=null
  — managed services (object storage, logging, monitoring, iam)    → group=null
  — external (on premises, internet, users)                        → group=null

═══════════════════════════════════════════════════════
INPUT SERVICES (from BOM + baseline injection):
═══════════════════════════════════════════════════════
[
{service_rows}
]

═══════════════════════════════════════════════════════
DEPLOYMENT HINTS (read from context — do not ask)
═══════════════════════════════════════════════════════
region_count:                   count explicit region mentions; default 1
availability_domains_per_region: 2 if "multi-AD" or "HA"; else 1
dr_enabled:                     true only if "DR" or "disaster recovery" stated
on_prem_connectivity:           "fastconnect" if DRG in BOM; "vpn" if VPN; "none" if neither

═══════════════════════════════════════════════════════
STEP 3 — DECLARE DATA-FLOW EDGES (minimal and purposeful)
═══════════════════════════════════════════════════════
OCI diagrams use ELEMENT POSITION to convey most connectivity.
Gateways straddling VCN edges and services grouped in subnets already
tell the reader how traffic flows — no line needed.

ONLY declare an explicit edge when the connection is NON-OBVIOUS from layout:
  ✓ Special network fabrics:    bare_metal ↔ bare_metal  (RDMA / RoCE)
  ✓ Storage mounts:             compute/bare_metal → file_storage  (NFS)
  ✓ Cross-boundary entry point: on_prem → drg  (FastConnect / VPN)
  ✓ Application data path:      compute → database  (SQL)
  ✓ Key API/service calls:      waf → load_balancer, functions → database

NEVER declare these — they are implied by gateway placement and need no line:
  ✗ internet → internet_gateway        (implied: IGW straddles VCN top)
  ✗ internet_gateway → waf/lb/bastion  (implied: IGW on VCN top edge)
  ✗ internet_gateway → region_box      (region_box is a container, not connectable)
  ✗ nat_gateway → internet             (implied: NAT on VCN top/right)
  ✗ service_gateway → any service      (implied: SGW on VCN right edge)
  ✗ drg → bastion / compute            (implied: DRG on VCN left edge)
  ✗ bastion → compute / bare_metal     (implied: co-located in same subnet)

CRITICAL: NEVER use region_box, vcn_box, or any subnet box ID as edge source or target.

Target: 2–5 edges. If the topology has fewer non-obvious connections, output [].

Example edges for this topology (adapt IDs to match INPUT SERVICES exactly):
{edge_examples}

Only declare edges between IDs that exist in INPUT SERVICES.
Use short protocol labels: "HTTPS/443", "SSH", "SQL/5432", "RDMA", "NFS", "Outbound", etc.

CRITICAL: edge source and target must be service icon IDs from INPUT SERVICES only.
NEVER use group/subnet IDs (e.g. bas_sub_box, cp_sub_box, worker_sub_box, storage_sub_box)
as edge source or target — those are layout containers, not connectable services.

═══════════════════════════════════════════════════════
OUTPUT FORMAT — LayoutIntent
═══════════════════════════════════════════════════════
{{
  "schema_version": "1.0",
  "deployment_hints": {{
    "region_count": 1,
    "availability_domains_per_region": 1,
    "dr_enabled": false,
    "on_prem_connectivity": "fastconnect"
  }},
  "groups": {example_groups_json},
  "placements": [
    {{"id": "<exact-id-from-input>", "oci_type": "<type>", "layer": "<layer>", "group": "<group-slug-or-null>"}},
    ...
  ],
  "edges": {edge_examples},
  "assumptions": [
    {{"id": "ha_mode", "statement": "Single AD assumed", "reason": "No HA signal in BOM", "risk": "low"}}
  ],
  "fixed_edges_policy": true
}}

OR — NeedClarification (ONLY if truly blocking):
{{"status": "need_clarification", "questions": [{{"id": "<id>", "question": "...", "blocking": true}}]}}

Allowed question IDs: regions.count, regions.mode, ha.ads, connectivity.onprem, dr.rpo_rto

RULES:
1. Every id from INPUT SERVICES must appear exactly once in placements.
2. Every group slug in placements must appear in the groups array.
3. Use exact id values from INPUT SERVICES — do not rename them.
4. edges must only reference ids that exist in INPUT SERVICES — NEVER group/subnet IDs like bas_sub_box.
5. Output ONLY valid JSON. No markdown, no prose, no code fences."""


def _parse_all_env_tabs(
    xlsx_path: "str | Path",
    wb,
    env_sheets: list[str],
    context: str = "",
) -> "list[ServiceItem]":
    """
    Parse all environment tabs and merge into one ServiceItem list.

    Primary tab: items keep their section notes (section "Primary" is renamed
    to the tab name so multi-env detection fires correctly).
    Secondary tabs: shared infra is dropped (already in primary); remaining
    items are re-ID'd with an env prefix and tagged with the tab name.

    This gives compile_intent_to_flat_spec enough notes diversity to trigger
    is_multi_env=True and route to _build_compartment_region.
    """
    import dataclasses as _dc

    all_items: list[ServiceItem] = []
    all_ids:   set[str]         = set()

    for tab_idx, tab_name in enumerate(env_sheets):
        is_primary  = (tab_idx == 0)
        tab_items   = parse_bom(xlsx_path, context=context, sheet_name=tab_name)
        tab_prefix  = f"{tab_name.lower()[:8].replace(' ', '_')}_"

        for item in tab_items:
            if is_primary:
                # Rename "Primary" → actual tab name so compiler detects it
                notes_val = tab_name if item.notes == "Primary" else item.notes
                all_items.append(_dc.replace(item, notes=notes_val))
                all_ids.add(item.id)
            else:
                # Secondary tabs: drop shared infra, baselines, on_prem duplicates
                if item.notes in {"best practice", "injected_baseline"}:
                    continue
                if item.oci_type in _SHARED_INFRA_TYPES:
                    continue
                if item.id == "on_prem":
                    continue
                # Env-prefix the ID; skip if already present (same oci_type, different qty)
                new_id = f"{tab_prefix}{item.oci_type.replace(' ', '_')}_1"
                if new_id in all_ids:
                    continue
                new_label = item.label.split("\n")[0].rstrip() + f"\n({tab_name})"
                all_items.append(_dc.replace(item, id=new_id, label=new_label, notes=tab_name))
                all_ids.add(new_id)

    return all_items


def bom_to_llm_input(
    xlsx_path: str | Path,
    context: str = "",
    questionnaire_text: str = "",
    notes_text: str = "",
) -> tuple[list[ServiceItem], str]:
    """Main entry point: parse BOM and return (items, llm_prompt).

    Uses build_layout_intent_prompt() by default (Option 1 architecture).
    context: optional free-text from an uploaded requirements/notes file.
    questionnaire_text: answers to a pre-flight questionnaire, if any.
    notes_text: meeting notes or other free-text, if any.

    Multi-tab BOMs (separate sheet per environment):
      Items from ALL env sheets are merged into one list, each tagged with its
      sheet name as `notes`.  This allows compile_intent_to_flat_spec to detect
      is_multi_env=True and route to the compartment layout path.
    """
    import openpyxl as _openpyxl
    wb = _openpyxl.load_workbook(xlsx_path, data_only=True)
    env_sheets = _find_bom_sheets(wb)

    if len(env_sheets) > 1:
        # Multi-tab BOM: parse all tabs and tag each item with its env name
        items       = _parse_all_env_tabs(xlsx_path, wb, env_sheets, context=context)
        env_summary = _build_env_summary(wb, env_sheets)
        full_context = f"{env_summary}\n\n{context}".strip() if context else env_summary
    else:
        items        = parse_bom(xlsx_path, context=context, sheet_name=env_sheets[0])
        full_context = context

    prompt = build_layout_intent_prompt(
        items,
        questionnaire_text=questionnaire_text,
        notes_text=notes_text,
        context=full_context,
    )
    return items, prompt
