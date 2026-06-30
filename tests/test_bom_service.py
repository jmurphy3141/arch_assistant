from __future__ import annotations

import time
from io import BytesIO

from openpyxl import load_workbook

from agent.bom_service import BomService, DEFAULT_PRICE_TABLE, CacheSnapshot


def _ready_service() -> BomService:
    svc = BomService()
    svc._cache = CacheSnapshot(
        pricing_table=dict(DEFAULT_PRICE_TABLE),
        shapes_text="fallback shapes",
        services_text="fallback services",
        refreshed_at=time.time(),
        source="test",
    )
    return svc


def test_bom_validation_enforces_unknown_sku_and_positive_price() -> None:
    svc = BomService()
    payload = {
        "currency": "USD",
        "line_items": [
            {
                "sku": "UNKNOWN",
                "description": "Bad row",
                "category": "compute",
                "quantity": 1,
                "unit_price": 10,
            },
            {
                "sku": "B94176",
                "description": "Compute",
                "category": "compute",
                "quantity": 1,
                "unit_price": 0,
            },
        ],
    }

    errors = svc.validate_final_payload(payload, dict(DEFAULT_PRICE_TABLE))
    assert any("unknown SKU" in err for err in errors)
    assert any("non-positive unit_price" in err for err in errors)


def test_bom_repair_adds_non_gpu_memory_split() -> None:
    svc = BomService()
    payload = {
        "currency": "USD",
        "line_items": [
            {
                "sku": "B94176",
                "description": "Compute E4 OCPU",
                "category": "compute",
                "quantity": 4,
                "unit_price": 0.05,
                "extended_price": 0.2,
            }
        ],
    }

    repaired, attempts, errors = svc._repair_until_valid(payload, dict(DEFAULT_PRICE_TABLE))
    assert attempts >= 1
    assert errors == []
    skus = {row["sku"] for row in repaired["line_items"]}
    assert "B94176" in skus
    assert "B94177" in skus


def test_bom_fast_path_normalizes_non_oci_provider_mentions() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        "Generate BOM for 8 OCPU, 128 GB RAM, 1 TB block storage with Hetzner load balancer and AWS object storage",
        dict(DEFAULT_PRICE_TABLE),
    )

    descriptions = " ".join(str(row["description"]) for row in payload["line_items"]).lower()
    notes = " ".join(str(row.get("notes") or "") for row in payload["line_items"]).lower()
    assert "hetzner" not in descriptions
    assert "aws" not in descriptions
    assert "hetzner" not in notes
    assert "aws" not in notes
    assert any(
        "oci-only bom enforced" in assumption.lower()
        for assumption in payload.get("assumptions", [])
    )


def test_bom_fast_path_honors_explicit_large_sizing() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        "Generate a BOM and XLSX for 48 OCPU, 768 GB RAM, and 42 TB block storage in us-ashburn-1.",
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B111129"]["quantity"] == 48.0
    assert by_sku["B111130"]["quantity"] == 768.0
    assert by_sku["B91961"]["quantity"] == 43008.0


def test_bom_fast_path_keeps_block_and_object_storage_quantities_distinct() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        "Create an OCI BOM with 500 GB Balanced Block Volume and 1 TB Standard Object Storage.",
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B91961"]["quantity"] == 500.0
    assert by_sku["B91962"]["quantity"] == 5000.0
    assert by_sku["B91628"]["quantity"] == 1024.0


def test_bom_fast_path_aggregates_word_form_server_count_and_preserves_region() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        "Create a BOM in us-phoenix-1 for three private VM.Standard.E5.Flex application "
        "servers at 4 OCPUs and 32 GB RAM each, with 2 TB Block Volume and 5 TB Object Storage.",
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert payload["region"] == "us-phoenix-1"
    assert by_sku["B97384"]["quantity"] == 12.0
    assert by_sku["B97384"]["instance_count"] == 3
    assert by_sku["B97385"]["quantity"] == 96.0
    assert by_sku["B91961"]["quantity"] == 2048.0
    assert by_sku["B91628"]["quantity"] == 5120.0


def test_bom_fast_path_ignores_memory_table_sizing_before_user_request() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        """[Archie Canonical Memory]
| Compute memory | 2048 GB |
| Block Volume | 1024 GB |
| Object Storage | 204.8 GB |
[End Archie Canonical Memory]
Create an OCI BOM for 2 servers, each with 2 OCPU and 16 GB memory, with 500 GB Balanced Block Volume and 1 TB Object Storage.""",
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B111130"]["quantity"] == 32.0
    assert by_sku["B91961"]["quantity"] == 500.0
    assert by_sku["B91628"]["quantity"] == 1024.0


def test_bom_fast_path_includes_waf_and_database_line_items() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        "Create an OCI XLSX bill of materials for an internet-facing 3-tier web app "
        "with WAF, public load balancer, private database layer, Object Storage, and Block Volume.",
        dict(DEFAULT_PRICE_TABLE),
    )

    skus = {row["sku"] for row in payload["line_items"]}
    assert "BWAF01" in skus
    assert "B99060" in skus


def test_structured_bom_inputs_drive_explicit_line_item_quantities() -> None:
    svc = _ready_service()

    result = svc.generate_from_inputs(
        inputs={
            "region": "af-johannesburg-1",
            "architecture_option": "OCI Dedicated VMware Solution",
            "compute": {"ocpu": 64, "gpu": False},
            "memory": {"gb": 1146.88},
            "storage": {"block_tb": 44, "object_tb": 5},
            "connectivity": {"internet_mbps": 100, "mpls": True, "sd_wan": True},
            "dr": {"rto_hours": 24, "cross_region_restore": True},
            "workloads": ["SQL Server", "Oracle databases", "Linux servers"],
            "os_mix": ["Linux", "Windows"],
            "output_format": "xlsx",
        },
        trace_id="trace-structured",
        model_id="test-bom",
    )

    assert result["type"] == "final"
    payload = result["bom_payload"]
    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B111129"]["quantity"] == 64.0
    assert by_sku["B111130"]["quantity"] == 1146.88
    assert by_sku["B91961"]["quantity"] == 45056.0
    assert by_sku["B91628"]["quantity"] == 5120.0
    assert by_sku["B111129"]["quantity"] != 4.0
    assert by_sku["B111130"]["quantity"] != 64.0
    assert by_sku["B91961"]["quantity"] != 1024.0
    assert payload["region"] == "af-johannesburg-1"
    assert payload["architecture_option"] == "OCI Dedicated VMware Solution"
    assert payload["workloads"] == ["SQL Server", "Oracle databases", "Linux servers"]
    assert payload["os_mix"] == ["Linux", "Windows"]


def test_structured_native_bom_includes_managed_oci_services() -> None:
    svc = _ready_service()

    result = svc.generate_from_inputs(
        inputs={
            "region": "af-johannesburg-1",
            "architecture_option": "OCI Native Services",
            "compute": {"ocpu": 64, "gpu": False},
            "memory": {"gb": 1147},
            "storage": {"block_tb": 44},
            "connectivity": {"internet_mbps": 100, "mpls": True, "sd_wan": False},
            "dr": {"rto_hours": 24, "cross_region_restore": True},
            "workloads": ["Oracle databases", "SQL Server", "file servers"],
            "target_services": [
                "VM.Standard.E5.Flex compute VMs",
                "Autonomous Database",
                "File Storage",
                "Block Volumes",
                "Object Storage backups/archive",
                "Load Balancer/WAF",
                "DRG/FastConnect/MPLS",
            ],
            "output_format": "xlsx",
        },
        trace_id="trace-native-structured",
        model_id="test-bom",
    )

    assert result["type"] == "final"
    payload = result["bom_payload"]
    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B97384"]["quantity"] == 64.0
    assert by_sku["B97385"]["quantity"] == 1147.0
    assert by_sku["B91961"]["quantity"] == 45056.0
    descriptions = " ".join(str(row["description"]) for row in payload["line_items"])
    notes = " ".join(str(row.get("notes") or "") for row in payload["line_items"])
    combined = f"{descriptions} {notes}"
    assert "VM.Standard.E5.Flex compute VMs" in combined
    assert "Autonomous Database" in combined
    assert "File Storage" in combined
    assert "Object Storage" in combined
    assert "Load Balancer" in combined
    assert "DRG/FastConnect/MPLS" in combined
    assert "OCI Native Services selected" in " ".join(payload["assumptions"])


def test_structured_bom_inputs_block_when_required_sizing_cannot_normalize() -> None:
    svc = _ready_service()

    result = svc.generate_from_inputs(
        inputs={
            "compute": {"ocpu": "sixty-four"},
            "memory": {"gb": 1146.88},
            "storage": {"block_tb": 44},
        },
        trace_id="trace-bad-inputs",
        model_id="test-bom",
    )

    assert result["type"] == "question"
    assert "compute.ocpu" in " ".join(result["normalization_blockers"])
    assert "bom_payload" not in result


def test_bom_fast_path_uses_kr1_migration_equivalent_table_values() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        """
Generate a BOM and XLSX for KR1 from the customer sizing table. Use the OCI
migration-equivalent sizing values, not the raw on-prem inventory values.

| Resource | Quantity | Specs |
|----------|----------|-------|
| VxRail CPU | 96 vCPU | equiv. ~48 OCPU for OCI migration target |
| VxRail RAM | 655 GB | OCI-equivalent RAM 768 GB for migration target |
| vSAN/HCI storage | 42.5 TB | usable capacity mapped to OCI Block Volume |
        """,
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B111129"]["quantity"] == 48.0
    assert by_sku["B111130"]["quantity"] == 768.0
    assert by_sku["B91961"]["quantity"] == 43520.0
    assert "source VxRail RAM 655 GB" in by_sku["B111130"]["notes"]
    assert "target OCI-equivalent RAM 768 GB" in by_sku["B111130"]["notes"]


def test_bom_validation_rejects_non_oci_provider_references_in_line_items() -> None:
    svc = BomService()
    payload = {
        "currency": "USD",
        "line_items": [
            {
                "sku": "B93030",
                "description": "Hetzner load balancer",
                "category": "network",
                "quantity": 1,
                "unit_price": 0.025,
                "notes": "Migrated from Hetzner",
            }
        ],
    }

    errors = svc.validate_final_payload(payload, dict(DEFAULT_PRICE_TABLE))
    assert any("non-oci provider" in err.lower() for err in errors)


def test_bom_fast_path_uses_markdown_table_quantities() -> None:
    svc = BomService()
    payload = svc._draft_bom_payload(
        """
| Category | Component | Specs/Details | Quantity |
|----------|-----------|---------------|----------|
| Compute (App Servers) | Ampere A1 Flex (Instance Pool/ASG) | 4 OCPU ARM, 24GB RAM, 200GB Block Vol, auto-scale min=3 | 3 |
| Load Balancer | Flexible Load Balancer (Standard Shape) | 10Mbps, L7 HTTP/S/HTTPS, path routing, health checks, WAF | 1 |
| Storage | Object Storage (Standard) | 250GB, 10TB egress free/yr | 1 |
        """,
        dict(DEFAULT_PRICE_TABLE),
    )

    by_sku = {row["sku"]: row for row in payload["line_items"]}
    assert by_sku["B93297"]["quantity"] == 12.0
    assert by_sku["B93298"]["quantity"] == 72.0
    assert by_sku["B91961"]["quantity"] == 600.0
    assert by_sku["B91962"]["quantity"] == 6000.0
    assert by_sku["B93030"]["quantity"] == 1.0
    assert by_sku["B91628"]["quantity"] == 250.0


def test_bom_pricing_parser_uses_oracle_usd_price_tiers() -> None:
    svc = BomService()
    table = svc._parse_pricing_payload(
        {
            "items": [
                {
                    "partNumber": "BTEST1",
                    "displayName": "OCI Test Service",
                    "metricName": "Unit Per Hour",
                    "currencyCodeLocalizations": [
                        {
                            "currencyCode": "USD",
                            "prices": [
                                {"model": "PAY_AS_YOU_GO", "value": 0},
                                {"model": "PAY_AS_YOU_GO", "rangeMin": 100, "value": 0.25},
                                {"model": "MONTHLY_FLEX", "value": 0.20},
                            ],
                        }
                    ],
                }
            ]
        }
    )

    assert table["BTEST1"]["unit_price"] == 0.25
    assert table["BTEST1"]["metric"] == "Unit Per Hour"
    assert table["BTEST1"]["source"] == "oracle_price_list_api"


def test_bom_xlsx_contains_scope_coverage_and_assumptions() -> None:
    payload = {
        "line_items": [
            {
                "sku": "B97384", "description": "E5 OCPU", "category": "compute",
                "metric": "OCPU Per Hour", "quantity": 4, "monthly_multiplier": 730,
                "unit_price": 0.03,
            }
        ],
        "scope_items": [
            {
                "canonical_service_id": "database.postgresql",
                "description": "PostgreSQL",
                "quantity": 1,
                "pricing_status": "authoritative_sku_required",
                "assumed_sizing": {"ocpu": 2, "memory_gb": 32, "storage_gb": 100},
                "assumptions": ["Assumed one PostgreSQL system for the POC."],
                "source": "selected_poc",
            }
        ],
    }

    workbook = load_workbook(BytesIO(BomService().generate_xlsx(payload)), data_only=False)

    assert workbook.sheetnames == ["BOM", "Scope Coverage"]
    scope = workbook["Scope Coverage"]
    assert scope["A2"].value == "database.postgresql"
    assert scope["B2"].value == "PostgreSQL"
    assert '"ocpu": 2' in scope["E2"].value
    assert "Assumed one PostgreSQL" in scope["F2"].value


def test_missing_public_price_stays_unpriced_and_out_of_total() -> None:
    service = BomService()
    table = service._parse_pricing_payload({"items": []})
    payload = service._normalize_payload({
        "line_items": [service._build_line(
            "BWAF01", 1, table, "security", "Selected WAF policy",
        )],
    })

    assert payload["line_items"][0]["unit_price"] is None
    assert payload["line_items"][0]["extended_price"] is None
    assert payload["line_items"][0]["pricing_status"] == "unpriced"
    assert payload["monthly_total"] == 0
    assert payload["pricing_status"] == "unpriced"


def test_strict_poc_inputs_preserve_selected_e5_and_exclude_unselected_services() -> None:
    service = _ready_service()
    result = service.generate_from_inputs(
        inputs={
            "region": "us-phoenix-1",
            "strict_scope": True,
            "scope": "poc",
            "target_service_ids": [
                "compute.vm.standard.e5.flex",
                "storage.object",
                "observability.logging",
            ],
            "compute": {"ocpu": 4, "instance_count": 1},
            "memory": {"gb": 32},
            "storage": {"object_tb": 1},
        },
        trace_id="strict-poc",
        model_id="test-bom",
    )

    payload = result["bom_payload"]
    assert payload["region"] == "us-phoenix-1"
    assert {row["sku"] for row in payload["line_items"]} == {"B97384", "B97385", "B91628"}
