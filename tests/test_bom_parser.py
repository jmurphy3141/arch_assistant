"""
tests/test_bom_parser.py
------------------------
Unit tests for agent.bom_parser.

Run: pytest tests/test_bom_parser.py
"""
import pytest
from pathlib import Path
from agent.bom_parser import (
    parse_bom, build_llm_prompt, bom_to_llm_input, ServiceItem,
    _normalize_desc, _infer_from_tokens, DESC_MAP,
)

FIXTURES = Path(__file__).parent / "fixtures"
SAMPLE_BOM = FIXTURES / "sample_bom.xlsx"


def _build_sample_bom(path: Path) -> None:
    """Generate a minimal BOM workbook covering key SKU_MAP entries."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    # Headers: SKU | Description | Quantity
    ws.append(["SKU", "Description", "Quantity"])
    ws.append(["B94176", "Compute VM Standard E4 Flex",    4])   # → compute / compute
    ws.append(["B99060", "Oracle Autonomous Database OCPU", 2])  # → database / data
    ws.append(["B93030", "Flexible Load Balancer",          1])  # → load balancer / ingress
    ws.append(["B88325", "Dynamic Routing Gateway",         2])  # → drg / ingress
    wb.save(path)


# Generate the fixture at collection time if it is absent
if not SAMPLE_BOM.exists():
    FIXTURES.mkdir(parents=True, exist_ok=True)
    _build_sample_bom(SAMPLE_BOM)


class TestBuildLlmPrompt:
    def test_returns_string(self):
        items = [
            ServiceItem(id="compute_1", oci_type="compute", label="Compute", layer="compute"),
            ServiceItem(id="on_prem",   oci_type="on premises", label="On-Premises", layer="external"),
        ]
        prompt = build_llm_prompt(items)
        assert isinstance(prompt, str)
        assert len(prompt) > 100

    def test_contains_service_ids(self):
        items = [ServiceItem(id="my_service", oci_type="compute", label="Compute", layer="compute")]
        prompt = build_llm_prompt(items)
        assert "my_service" in prompt

    def test_context_injected(self):
        items = [ServiceItem(id="c1", oci_type="compute", label="C", layer="compute")]
        prompt = build_llm_prompt(items, context="6 regions, active-passive HA")
        assert "6 regions" in prompt

    def test_no_context_no_injection(self):
        items = [ServiceItem(id="c1", oci_type="compute", label="C", layer="compute")]
        prompt = build_llm_prompt(items, context="")
        assert "ADDITIONAL CONTEXT" not in prompt

    def test_clarification_rule_present(self):
        items = [ServiceItem(id="c1", oci_type="compute", label="C", layer="compute")]
        prompt = build_llm_prompt(items)
        assert "need_clarification" in prompt


class TestParseBom:
    def test_returns_list(self):
        items = parse_bom(SAMPLE_BOM)
        assert isinstance(items, list)
        assert len(items) > 0

    def test_on_prem_always_present(self):
        items = parse_bom(SAMPLE_BOM)
        types = [i.oci_type for i in items]
        assert "on premises" in types

    def test_best_practice_services_added(self):
        items = parse_bom(SAMPLE_BOM)
        types = {i.oci_type for i in items}
        assert "internet gateway" in types
        assert "nat gateway" in types
        assert "service gateway" in types
        # Baseline injections
        assert "internet" in types
        assert "bastion" in types

    def test_internet_injected_deterministically(self):
        items = parse_bom(SAMPLE_BOM)
        ids = {i.id for i in items}
        types = {i.oci_type for i in items}
        assert "internet" in ids
        assert "internet" in types
        internet = next(i for i in items if i.id == "internet")
        assert internet.layer == "external"
        assert internet.notes == "injected_baseline"

    def test_bastion_injected_deterministically(self):
        items = parse_bom(SAMPLE_BOM)
        ids = {i.id for i in items}
        types = {i.oci_type for i in items}
        assert "bastion_1" in ids
        assert "bastion" in types
        bastion = next(i for i in items if i.id == "bastion_1")
        assert bastion.layer == "ingress"
        assert bastion.notes == "injected_baseline"

    def test_suppress_internet_via_context(self):
        items = parse_bom(SAMPLE_BOM, context="NO_INTERNET_ENDPOINT=true")
        types = {i.oci_type for i in items}
        assert "internet" not in types

    def test_suppress_bastion_via_context(self):
        items = parse_bom(SAMPLE_BOM, context="NO_BASTION=true")
        types = {i.oci_type for i in items}
        assert "bastion" not in types

    def test_no_duplicate_internet_injection(self):
        items = parse_bom(SAMPLE_BOM)
        internet_items = [i for i in items if i.oci_type == "internet"]
        assert len(internet_items) == 1

    def test_no_duplicate_bastion_injection(self):
        items = parse_bom(SAMPLE_BOM)
        bastion_items = [i for i in items if i.oci_type == "bastion"]
        assert len(bastion_items) == 1

    def test_all_items_have_required_fields(self):
        items = parse_bom(SAMPLE_BOM)
        for item in items:
            assert item.id
            assert item.oci_type
            assert item.label
            assert item.layer in ("external", "ingress", "compute", "async", "data")

    def test_non_bom_sheet_name_accepted(self, tmp_path):
        """parse_bom must work when the sheet is not named 'BOM'."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Services"          # not "BOM"
        ws.append(["SKU", "Description", "Quantity"])
        ws.append(["B94176", "Compute VM Standard E4 Flex", 2])
        ws.append(["B99060", "Oracle Autonomous Database",  1])
        path = tmp_path / "custom_name.xlsx"
        wb.save(path)

        items = parse_bom(path)
        types = {i.oci_type for i in items}
        assert "compute" in types
        assert "database" in types


class TestDescNormalization:
    """_normalize_desc strips Oracle/OCI branding prefixes."""

    def test_strips_oracle_cloud_infrastructure(self):
        assert _normalize_desc("oracle cloud infrastructure mysql heatwave") == "mysql heatwave"

    def test_strips_oracle_cloud(self):
        assert _normalize_desc("oracle cloud analytics cloud") == "analytics cloud"

    def test_strips_oci_prefix(self):
        assert _normalize_desc("oci functions - invocations") == "functions - invocations"

    def test_strips_oracle_prefix(self):
        assert _normalize_desc("oracle goldengate - ocpu") == "goldengate - ocpu"

    def test_no_prefix_unchanged(self):
        assert _normalize_desc("mysql heatwave - node") == "mysql heatwave - node"


class TestTokenInference:
    """_infer_from_tokens handles OCI billing rows not in DESC_MAP."""

    def test_ecpu_row_is_database(self):
        assert _infer_from_tokens("autonomous database - ecpu compute per hour") == ("database", "data")

    def test_ocpu_row_is_compute(self):
        assert _infer_from_tokens("vm.standard3.flex - ocpu per hour") == ("compute", "compute")

    def test_ocpu_autonomous_is_database(self):
        assert _infer_from_tokens("autonomous transaction processing - ocpu per hour") == ("database", "data")

    def test_bandwidth_is_skipped(self):
        assert _infer_from_tokens("outbound data transfer - bandwidth") == (None, None)

    def test_storage_gb_is_skipped(self):
        assert _infer_from_tokens("block volume storage - gb per month") == (None, None)

    def test_unknown_returns_none(self):
        # A genuinely novel description with no billing tokens → caller logs warning
        assert _infer_from_tokens("oracle resource manager bundle") is None


class TestDescMapCoversKnownProducts:
    """Spot-check that common OCI products resolve via DESC_MAP (tier-2 lookup)."""

    def _resolve(self, desc: str):
        from agent.bom_parser import _normalize_desc, DESC_MAP
        for raw in (desc, _normalize_desc(desc)):
            for key, val in DESC_MAP.items():
                if key in raw:
                    return val
        return None

    @pytest.mark.parametrize("desc,expected_type", [
        ("oracle cloud infrastructure mysql heatwave - node ocpu", "database"),
        ("oracle cloud infrastructure autonomous database serverless - ecpu", "database"),
        ("oracle cloud infrastructure goldengate - ocpu per hour",  "compute"),
        ("oracle cloud infrastructure data science - notebook session ocpu", "compute"),
        ("oracle cloud infrastructure analytics cloud - professional ocpu", "compute"),
        ("oracle cloud infrastructure big data - master node ocpu", "compute"),
        ("oracle cloud infrastructure opensearch - ocpu", "database"),
        ("oracle cloud infrastructure cloud guard - target",  "monitoring"),
        ("oracle cloud infrastructure key management - key version", "vault"),
        ("oracle cloud infrastructure streaming - data read api", "queue"),
        ("oracle cloud infrastructure functions - invocations million", "functions"),
        ("oracle cloud infrastructure object storage - storage", "object storage"),
        ("oracle cloud infrastructure file storage - metered iops", "file storage"),
        ("oracle cloud infrastructure network load balancer - lcus", "load balancer"),
        ("oracle cloud infrastructure web application firewall - request", "waf"),
    ])
    def test_known_product(self, desc, expected_type):
        result = self._resolve(desc)
        assert result is not None, f"No match for: {desc!r}"
        oci_type, _ = result
        assert oci_type == expected_type, f"{desc!r} → {oci_type!r}, want {expected_type!r}"


class TestUnknownSkuEndToEnd:
    """parse_bom must handle rows with novel SKUs via description fallback."""

    def test_novel_sku_resolved_via_desc(self, tmp_path):
        """A SKU not in SKU_MAP resolves via DESC_MAP description keyword."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["SKU", "Description", "Quantity"])
        ws.append(["ZZZUNKNOWN", "Oracle Cloud Infrastructure MySQL HeatWave - Node OCPU", 2])
        path = tmp_path / "novel.xlsx"
        wb.save(path)

        items = parse_bom(path)
        types = {i.oci_type for i in items}
        assert "database" in types, "MySQL HeatWave should resolve to 'database'"

    def test_ecpu_row_resolved_via_token_inference(self, tmp_path):
        """A row with 'ECPU' in a novel description infers database via token logic."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["SKU", "Description", "Quantity"])
        ws.append(["BFUTURE1", "Oracle Future Database Service - ECPU per hour", 4])
        path = tmp_path / "ecpu.xlsx"
        wb.save(path)

        items = parse_bom(path)
        types = {i.oci_type for i in items}
        assert "database" in types, "ECPU row should infer database"

    def test_ocpu_row_resolved_via_token_inference(self, tmp_path):
        """A row with only 'OCPU' in a novel description infers compute via token logic."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["SKU", "Description", "Quantity"])
        ws.append(["BFUTURE2", "Oracle Cloud Infrastructure Future.Flex - OCPU per hour", 8])
        path = tmp_path / "ocpu.xlsx"
        wb.save(path)

        items = parse_bom(path)
        types = {i.oci_type for i in items}
        assert "compute" in types, "Novel OCPU row should infer compute"

    def test_billing_row_silently_skipped(self, tmp_path):
        """Bandwidth / data-transfer rows produce no node."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["SKU", "Description", "Quantity"])
        ws.append(["BBANDWIDTH", "Outbound Data Transfer - Bandwidth per GB", 100])
        path = tmp_path / "bandwidth.xlsx"
        wb.save(path)

        items = parse_bom(path)
        # Only baseline items should be present — no "compute" or "database" from the BOM row
        baseline_types = {"on premises", "internet gateway", "nat gateway", "service gateway",
                          "waf", "bastion", "logging", "monitoring", "iam", "vault", "internet"}
        bom_derived = [i for i in items if i.oci_type not in baseline_types]
        assert not bom_derived, f"Bandwidth row should produce no node, got: {[i.oci_type for i in bom_derived]}"
