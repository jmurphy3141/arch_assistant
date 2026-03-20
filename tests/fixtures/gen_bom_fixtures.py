"""
tests/fixtures/gen_bom_fixtures.py
------------------------------------
Generates BOM Excel fixtures for the three Calypso/Capital Markets live LLM test
scenarios. Run once to create the .xlsx files.

Usage:
    python tests/fixtures/gen_bom_fixtures.py

Output:
    tests/fixtures/calypso_s1_full.xlsx
    tests/fixtures/calypso_s2_partial.xlsx
    tests/fixtures/calypso_s3_minimal.xlsx

BOM format expected by agent/bom_parser.parse_bom():
  Sheet "BOM"   — columns: SKU, Description, Quantity, Notes
  Sheet "Input" — columns: Service, vCPU Count, Storage (GB), Notes
                  rows:    EC2 (app servers), Postgres RDS (db)
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

FIXTURES_DIR = Path(__file__).parent


def _header_style():
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(fill_type="solid", fgColor="312D2A")  # OCI dark
    alignment = Alignment(horizontal="center")
    return font, fill, alignment


def _write_bom(path: Path, bom_rows: list[tuple], input_rows: list[tuple]) -> None:
    """
    bom_rows:   list of (SKU, Description, Quantity, Notes)
    input_rows: list of (Service, vCPU Count, Storage (GB), Notes)
    """
    wb = openpyxl.Workbook()

    # ── BOM sheet ──────────────────────────────────────────────────────────────
    ws_bom = wb.active
    ws_bom.title = "BOM"

    bom_headers = ["SKU", "Description", "Quantity", "Notes"]
    font, fill, align = _header_style()
    for col, hdr in enumerate(bom_headers, 1):
        cell = ws_bom.cell(row=1, column=col, value=hdr)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    for r, row in enumerate(bom_rows, 2):
        for c, val in enumerate(row, 1):
            ws_bom.cell(row=r, column=c, value=val)

    ws_bom.column_dimensions["A"].width = 12
    ws_bom.column_dimensions["B"].width = 40
    ws_bom.column_dimensions["C"].width = 12
    ws_bom.column_dimensions["D"].width = 50

    # ── Input sheet ────────────────────────────────────────────────────────────
    ws_inp = wb.create_sheet("Input")
    inp_headers = ["Service", "vCPU Count", "Storage (GB)", "Notes"]
    for col, hdr in enumerate(inp_headers, 1):
        cell = ws_inp.cell(row=1, column=col, value=hdr)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

    for r, row in enumerate(input_rows, 2):
        for c, val in enumerate(row, 1):
            ws_inp.cell(row=r, column=c, value=val)

    ws_inp.column_dimensions["A"].width = 20
    ws_inp.column_dimensions["B"].width = 14
    ws_inp.column_dimensions["C"].width = 16
    ws_inp.column_dimensions["D"].width = 40

    wb.save(path)
    print(f"  Written: {path}")


# ── OCI SKUs used in BOM parser ────────────────────────────────────────────────
# B94176 → compute        B99060 → database       B91628 → object storage
# B93030 → load balancer  B88325 → drg            B92072 → api gateway
# B95697 → queue          B90618 → functions
# (bastion / network firewall come via DESC_MAP keyword match)

def gen_scenario1():
    """Full information — all sizing specified."""
    bom_rows = [
        # SKU            Description                                Qty    Notes
        ("B94176",  "Compute — VM.Standard3.Flex (App Tier)",       16,   "8 nodes per AD × 2 ADs; 16 OCPU 256 GB each"),
        ("B99060",  "Oracle DB EE — RAC (Exadata X9M config)",       4,   "2-node RAC per AD × 2 ADs; 32 OCPU 512 GB each"),
        ("B91628",  "Object Storage — market data & RMAN backup",    10,   "10 TB total (TB units)"),
        ("B93030",  "Load Balancer — public (WAF + HTTP/HTTPS)",      2,   "1 per AD for external counterparty portal"),
        ("B93030",  "Network Load Balancer — private (FastConnect)",  2,   "1 per AD for internal/DRG traffic"),
        ("B88325",  "DRG — FastConnect 10 Gbps × 2 ports",           2,   "Redundant FastConnect; Chicago + London offices"),
        ("B92072",  "API Gateway — trading counterparty APIs",        1,   "Regional service; REST + FIX protocol adapters"),
        ("B95697",  "OCI Streaming (Kafka-compatible)",               1,   "Bloomberg market data feed 500k msgs/sec peak"),
        (None,      "Bastion Host — admin/ops SSH access",            1,   "Public subnet; key-pair auth only"),
        (None,      "Network Firewall — ingress inspection",          1,   "IDS/IPS on public ingress path"),
    ]
    input_rows = [
        # Service         vCPU Count   Storage (GB)  Notes
        ("EC2",              256,        None,        "8 nodes × 16 OCPU × 2 (hyperthreading factor) × 2 ADs"),
        ("Postgres RDS",     128,        51200,       "32 OCPU × 2 ADs; 50 TB NVMe block per AD = 51200 GB"),
    ]
    _write_bom(FIXTURES_DIR / "calypso_s1_full.xlsx", bom_rows, input_rows)


def gen_scenario2():
    """Partial information — sizing gaps; architect knows services but not counts."""
    bom_rows = [
        # SKU            Description                               Qty    Notes
        ("B94176",  "Compute — App Tier servers",                  None, "Count not specified; several app servers"),
        ("B99060",  "Oracle Database EE — edition TBD",            None, "RAC likely; exact sizing not confirmed"),
        ("B91628",  "Object Storage — backup",                     None, "Size TBD; backup destination"),
        ("B93030",  "Load Balancer",                               2,    "Public + private LB required"),
        ("B88325",  "DRG — FastConnect",                           1,    "Single FastConnect; bandwidth not specified"),
        (None,      "Bastion Host",                                1,    "Required for admin access"),
    ]
    input_rows = [
        # Service         vCPU Count   Storage (GB)  Notes
        ("EC2",           None,        None,        "Sizing not provided — use OCI recommended defaults"),
        ("Postgres RDS",  None,        None,        "Sizing not provided — Oracle edition and OCPU TBD"),
    ]
    _write_bom(FIXTURES_DIR / "calypso_s2_partial.xlsx", bom_rows, input_rows)


def gen_scenario3():
    """Minimal information — only know it is a capital markets trading platform."""
    bom_rows = [
        # SKU    Description                          Qty    Notes
        ("B94176",  "Compute",                        None, "Trading platform — sizing unknown"),
        ("B99060",  "Database",                       None, "Capital markets; likely Oracle RAC"),
        ("B88325",  "DRG",                            None, "On-premises connectivity required"),
    ]
    input_rows = [
        # Service         vCPU Count   Storage (GB)  Notes
        ("EC2",           None,        None,        "No sizing info provided"),
        ("Postgres RDS",  None,        None,        "No sizing info provided"),
    ]
    _write_bom(FIXTURES_DIR / "calypso_s3_minimal.xlsx", bom_rows, input_rows)


if __name__ == "__main__":
    print("Generating BOM fixtures...")
    gen_scenario1()
    gen_scenario2()
    gen_scenario3()
    print("Done.")
